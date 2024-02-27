from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.workbook import Workbook

from funcoes.get.get_linhas_json import (get_coluna_fator, get_coluna_index,
                                         get_linha_fator, get_planilha_fator,
                                         get_valor_bdi_formatado)


def adicionar_bdi(workbook: Workbook, dados):
    # Obter valores
    planilha_fator = get_planilha_fator(dados)
    valor_bdi_formatado = get_valor_bdi_formatado(dados)
    linha = get_linha_fator(dados)
    colunaString = get_coluna_fator(dados)
    colunaNumber = get_coluna_index(colunaString)

    # Selecionar a planilha
    sheet_resumo = workbook[planilha_fator]

    # Selecionar a célula
    cell = sheet_resumo.cell(
        row=linha, column=colunaNumber)

    # Definir o valor da célula como "FATOR"
    cell.value = "BDI"

    # Selecionar a célula
    cell_value = sheet_resumo.cell(
        row=linha, column=colunaNumber+1)

    # Definir o valor da célula
    cell_value.value = valor_bdi_formatado

    # Adicionar a definição de nome "FATOR" referenciando a célula
    ref = f"{planilha_fator}!${get_column_letter(colunaNumber+1)}${linha}"
    definicao_nome = DefinedName(name="BDI", attr_text=ref)
    workbook.defined_names.add(definicao_nome)
