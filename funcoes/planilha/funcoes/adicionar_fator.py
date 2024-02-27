from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.workbook import Workbook

from funcoes.get.get_linhas_json import (get_coluna_fator, get_linha_fator,
                                         get_planilha_fator)


def adicionar_fator(workbook: Workbook, dados):
    # Obter valores
    planilha_fator = get_planilha_fator(dados)
    linha = get_linha_fator(dados)
    colunaString = get_coluna_fator(dados)
    colunaNumber = column_index_from_string(colunaString)

    # Selecionar a planilha
    sheet_resumo = workbook[planilha_fator]

    # Selecionar a célula
    cell = sheet_resumo.cell(
        row=linha, column=colunaNumber)

    # Definir o valor da célula como "FATOR"
    cell.value = "FATOR"

    # Selecionar a célula
    cell_value = sheet_resumo.cell(
        row=linha, column=colunaNumber+1)

    # Definir o valor da célula como "1"
    cell_value.value = 1  # Se for um número, não é necessário aspas

    # Adicionar a definição de nome "FATOR" referenciando a célula
    ref = f"{planilha_fator}!${get_column_letter(colunaNumber+1)}${linha}"
    definicao_nome = DefinedName(name="FATOR", attr_text=ref)
    workbook.defined_names.add(definicao_nome)
