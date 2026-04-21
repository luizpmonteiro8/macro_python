from openpyxl.utils import column_index_from_string
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.hyperlink import Hyperlink
from funcoes.common.buscar_palavras import (
    buscar_palavra_contem,
    buscar_palavra_com_linha,
    buscar_palavra_com_linha_iniciando,
)
from funcoes.get.get_linhas_json import (
    get_planilha_comp,
    get_descricao_comp,
    get_planilha_aux,
    get_descricao_aux,
    get_item_descricao_comp_aux,
    get_preco_unitario_comp,
    get_coluna_totais_aux,
    get_valor_totais_aux,
    get_valor_string,
)


def criar_hiperlinks_auxiliares(workbook, dados, todos_item):
    """
    Para cada item em COMPOSICOES:
    - Busca na COMPOSICOES AUXILIARES usando lógica robusta (código, prefixos, descrição)
    - Cria hyperlink na coluna descrição -> item na AUXILIARES
    - Cria hyperlink interno no item da AUXILIARES -> VALOR: da seção
    - Cria fórmula em PREÇO UNITÁRIO -> VALOR: do item
    """
    print(">>> Criando hyperlinks para itens auxiliares...")

    dados_itens = todos_item
    if isinstance(todos_item, list) and len(todos_item) > 0:
        dados_itens = todos_item[0]

    planilha_comp = get_planilha_comp(dados)
    planilha_aux = get_planilha_aux(dados)
    col_item = get_descricao_comp(dados)
    col_desc_link = get_item_descricao_comp_aux(dados)
    col_desc_aux = get_descricao_aux(dados)
    col_preco_comp = get_preco_unitario_comp(dados)
    col_valor_aux = get_valor_totais_aux(dados)
    col_totais_aux = get_coluna_totais_aux(dados)
    valor_str = get_valor_string(dados)

    sheet_comp = workbook[planilha_comp]
    sheet_aux = workbook[planilha_aux]
    sheet_comp_max = sheet_comp.max_row + 1
    sheet_aux_max = sheet_aux.max_row + 1

    col_item_idx = ord(col_item.upper()) - ord("A") + 1
    col_desc_link_idx = ord(col_desc_link.upper()) - ord("A") + 1
    col_preco_idx = ord(col_preco_comp.upper()) - ord("A") + 1
    col_desc_aux_idx = ord(col_desc_aux.upper()) - ord("A") + 1
    col_valor_aux_idx = ord(col_valor_aux.upper()) - ord("A") + 1
    col_totais_aux_idx = ord(col_totais_aux.upper()) - ord("A") + 1

    hyperlinks_criados = 0
    formulas_criadas = 0
    ultima_linha_aux = sheet_aux_max
    ultima_busca = 1

    for x in range(1, sheet_comp_max):
        cod = sheet_comp.cell(row=x, column=col_item_idx).value
        desc = sheet_comp.cell(row=x, column=col_item_idx).value
        if cod is None:
            continue

        cod = str(cod).strip()
        desc_str = str(desc).strip() if desc else ""

        desc_upper = desc_str.upper()
        if any(
            x in desc_upper
            for x in [
                "COEFICIENTE",
                "PREÇO UNITÁRIO",
                "FONTE",
                "UNID",
                "TOTAL",
                "VALOR",
            ]
        ):
            continue

        if not cod:
            continue

        print(f"busca item {cod} {desc_str[:30]} na linha {x}")

        linha_ini = -1

        if linha_ini == -1:
            linha_ini = buscar_palavra_com_linha_iniciando(
                sheet_aux, col_desc_aux, cod, ultima_busca, ultima_linha_aux
            )

        if linha_ini == -1:
            linha_ini = buscar_palavra_com_linha_iniciando(
                sheet_aux, col_desc_aux, cod, 1, ultima_linha_aux
            )

        if linha_ini == -1:
            linha_ini = buscar_palavra_contem(
                sheet_aux, col_desc_aux, cod, 1, ultima_linha_aux
            )

        if linha_ini == -1 and len(cod) >= 5:
            cod_prefix = cod[:5]
            linha_ini = buscar_palavra_contem(
                sheet_aux, col_desc_aux, cod_prefix, 1, ultima_linha_aux
            )

        if linha_ini == -1 and len(cod) >= 4:
            cod_prefix = cod[:4]
            linha_ini = buscar_palavra_contem(
                sheet_aux, col_desc_aux, cod_prefix, 1, ultima_linha_aux
            )

        if linha_ini == -1 and desc_str:
            linha_ini = buscar_palavra_contem(
                sheet_aux, col_desc_aux, desc_str, 1, ultima_linha_aux
            )

        if linha_ini == -1 and desc_str:
            palavras = desc_str.split()
            for palavra in palavras:
                if len(palavra) > 5 and not palavra.isdigit():
                    linha_ini = buscar_palavra_contem(
                        sheet_aux, col_desc_aux, palavra, 1, ultima_linha_aux
                    )
                    if linha_ini > -1:
                        break

        if linha_ini == -1:
            print(f"⚠️ Item não encontrado: {cod}")
            continue

        print(f">> L{x} -> AUXILIARES L{linha_ini}")

        ultima_busca = 1
        linha_fim = buscar_palavra_com_linha(
            sheet_aux, col_totais_aux, valor_str, linha_ini, ultima_linha_aux
        )

        if linha_fim == -1:
            print(f"⚠️ VALOR: não encontrado para: {cod}")
            continue

        # Hyperlink em COMPOSICOES col descrição -> AUXILIARES item Não alterar esta funcionando
        cell_link = sheet_comp.cell(row=x, column=col_desc_link_idx)
        if not isinstance(cell_link, MergedCell):
            location = f"{planilha_aux}!{col_desc_aux}{linha_ini}"
            cell_link.hyperlink = f"#'{planilha_aux}'!{col_desc_aux}{linha_ini}"
            hyperlinks_criados += 1

        # Hyperlink interno em AUXILIARES -> linha do item encontrado (descrição)
        cell_aux_link = sheet_aux.cell(row=int(linha_ini), column=int(col_desc_aux_idx))
        if not isinstance(cell_aux_link, MergedCell):
            location = f"{planilha_aux}!{col_desc_aux}{int(linha_ini)}"
            cell_aux_link.hyperlink = Hyperlink(
                ref=f"{col_desc_aux}{int(linha_ini)}",
                location=location,
                display=cell_aux_link.value,
            )

        # Fórmula em PREÇO UNITÁRIO
        cell_preco = sheet_comp.cell(row=x, column=col_preco_idx)
        if isinstance(cell_preco, MergedCell):
            continue
        if not (
            cell_preco.value
            and isinstance(cell_preco.value, str)
            and cell_preco.value.startswith("=")
        ):
            cell_preco.value = f"='{planilha_aux}'!{col_valor_aux}{linha_fim}"
            formulas_criadas += 1

    print(f">> Hyperlinks criados: {hyperlinks_criados}")
    print(f">> Fórmulas criadas: {formulas_criadas}")

    return hyperlinks_criados, formulas_criadas
