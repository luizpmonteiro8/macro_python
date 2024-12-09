from copy import copy

from openpyxl import Workbook
from openpyxl.utils import column_index_from_string

from funcoes.common.copiar_coluna import copiar_coluna
from funcoes.get.get_linhas_json import (get_planilha_orcamentaria,
                                         get_planilha_preco_total,
                                         get_planilha_preco_unitario,
                                         get_planilha_preco_unitario_copiar,
                                         get_planilha_quantidade)


def copiar_coluna_planilha(sheet, dados):
    # Obter informações de coluna do JSON
    coluna_origem = get_planilha_preco_unitario(dados)
    coluna_destino = get_planilha_preco_unitario_copiar(dados)

    copiar_coluna(sheet, coluna_origem, coluna_destino)


def copiar_estilo(origem, destino):
    destino.font = copy(origem.font)
    destino.border = copy(origem.border)
    destino.fill = copy(origem.fill)
    destino.number_format = copy(origem.number_format)
    destino.protection = copy(origem.protection)
    destino.alignment = copy(origem.alignment)


def formula_planilha(woorBook: Workbook, linhaIni, linhaFin, dados):
    # Obter informações de coluna do JSON
    planilha = get_planilha_orcamentaria(dados)

    quantidade_string = get_planilha_quantidade(dados)
    preco_unitario_string = get_planilha_preco_unitario(dados)

    coluna_preco_total = get_planilha_preco_total(dados)
    coluna_preco_total_number = column_index_from_string(coluna_preco_total)

    coluna_value_string = get_planilha_preco_unitario_copiar(dados)
    coluna_value_number = column_index_from_string(coluna_value_string)

    ws = woorBook[planilha]

    # Iterar sobre as linhas
    for x in range(linhaIni, linhaFin):
        # Verificar se o valor na coluna 11 não está vazio ou seja não é total
        if ws.cell(row=x, column=coluna_value_number).value is not None:
            # Calcular e atribuir o valor na coluna
            ws.cell(row=x, column=coluna_value_number+1).value = (
                f"={preco_unitario_string}{x}-"
                f"{coluna_value_string}{x}"
            )

            # Copiar o estilo da célula de quantidade para preço total
            copiar_estilo(ws.cell(row=x, column=column_index_from_string(
                quantidade_string)), ws.cell(row=x,
                                             column=coluna_preco_total_number))

            # Calcular e atribuir o valor na coluna de preço total
            #ws.cell(row=x, column=coluna_preco_total_number).value = (
            #    f"=ROUND({quantidade_string}{x}*{preco_unitario_string}{x}, 2)"
            #)
