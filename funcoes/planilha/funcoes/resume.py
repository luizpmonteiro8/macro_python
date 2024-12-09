from openpyxl import Workbook

from funcoes.common.buscar_palavras import buscar_palavra_com_linha
from funcoes.get.get_linhas_json import get_coluna_final, get_planilha_orcamentaria, get_planilha_preco_total

# pega nome no resumo buscao na planilha orcamentaria, adiciona forma da porcentagem adiciona os totais
def resumo_totais(workbook: Workbook, dados):
    sheet_name = get_planilha_orcamentaria(dados)
    coluna_total = get_planilha_preco_total(dados)

    ws = workbook["RESUMO"]
    wsOrcamento = workbook[sheet_name]
    orcamentoLinhaFinal = wsOrcamento.max_row+1
    resumoLinhaFinal = ws.max_row+1   

    linha_resumo_total = buscar_palavra_com_linha(
               ws, "C", "VALOR TOTAL:",
               1, resumoLinhaFinal) 


    for x in range(1, ws.max_row + 1):
        if ws.cell(row=x, column=2).value is not None:
           nome = ws.cell(row=x, column=2).value
           linha_inicial = buscar_palavra_com_linha(
               wsOrcamento, "B", nome,
               1, orcamentoLinhaFinal)   
           if (linha_inicial != -1):
                ws[f'{"D"}{x}'].value = (
                   f'=\'{sheet_name}\'!{coluna_total}{linha_inicial}')
                ws[f'{"E"}{x}'].value = (
                   f'=(D{x}/$D${linha_resumo_total})*100')

    #totais
    linha_orcamento_total = buscar_palavra_com_linha(
           wsOrcamento, get_coluna_final(dados), "VALOR TOTAL:",
           1, orcamentoLinhaFinal) 
    if (linha_orcamento_total != -1):
        ws[f'{"D"}{linha_resumo_total}'].value = (
            f'=\'{sheet_name}\'!{coluna_total}{linha_orcamento_total}')
        ws[f'{"D"}{linha_resumo_total-1}'].value = (
            f'=\'{sheet_name}\'!{coluna_total}{linha_orcamento_total-1}')
        ws[f'{"D"}{linha_resumo_total-2}'].value = (
            f'=\'{sheet_name}\'!{coluna_total}{linha_orcamento_total-2}')
        

        