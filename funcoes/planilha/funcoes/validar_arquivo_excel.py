import json
import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl.utils import column_index_from_string

from funcoes.common.buscar_palavras import buscar_palavra
from funcoes.get.get_linhas_json import *


CAMINHO_JSON = "config/valores_colunas.json"


def janela_corrigir_valor(titulo, mensagem, instrucao, valor_atual, valor_default=None):
    """
    Mostra janela para usuário corrigir um valor.

    Args:
        titulo: Título da janela
        mensagem: O que está errado (leigo)
        instrucao: Passo a passo para encontrar o valor no Excel
        valor_atual: Valor atual no JSON (para referência)
        valor_default: Valor sugerido como padrão

    Returns:
        tuple: (True, novo_valor) se confirmado, (False, None) se cancelado
    """
    resultado = {"confirmado": False, "valor": None}

    def on_confirmar():
        resultado["confirmado"] = True
        resultado["valor"] = entry.get()
        janela.destroy()

    def on_cancelar():
        janela.destroy()

    janela = tk.Toplevel()
    janela.title(titulo)
    janela.geometry("600x400")
    janela.resizable(False, False)
    janela.grab_set()
    janela.focus_force()
    janela.transient()

    mainframe = tk.Frame(janela, padx=20, pady=20)
    mainframe.pack(fill=tk.BOTH, expand=True)

    tk.Label(
        mainframe,
        text=mensagem,
        font=("Arial", 12, "bold"),
        wraplength=550,
        justify=tk.LEFT,
    ).pack(anchor=tk.W)

    tk.Label(mainframe, text="", font=("Arial", 8)).pack()

    tk.Label(mainframe, text="INSTRUÇÕES:", font=("Arial", 10, "bold")).pack(
        anchor=tk.W
    )
    tk.Label(
        mainframe, text=instrucao, font=("Arial", 10), justify=tk.LEFT, wraplength=550
    ).pack(anchor=tk.W)

    tk.Label(mainframe, text="", font=("Arial", 8)).pack()

    tk.Label(
        mainframe,
        text=f"Valor atual no sistema: '{valor_atual}'",
        font=("Arial", 9),
        fg="gray",
    ).pack(anchor=tk.W)

    tk.Label(mainframe, text="Digite o novo valor:", font=("Arial", 10, "bold")).pack(
        anchor=tk.W, pady=(10, 5)
    )

    entry = tk.Entry(mainframe, width=50, font=("Arial", 11))
    if valor_default:
        entry.insert(0, valor_default)
    elif valor_atual:
        entry.insert(0, valor_atual)
    entry.pack(fill=tk.X, pady=(0, 10))
    entry.focus()

    frame_botoes = tk.Frame(mainframe)
    frame_botoes.pack(fill=tk.X)

    btn_cancelar = tk.Button(
        frame_botoes, text="Cancelar", font=("Arial", 10), width=15, command=on_cancelar
    )
    btn_cancelar.pack(side=tk.RIGHT, padx=(10, 0))

    btn_confirmar = tk.Button(
        frame_botoes,
        text="Corrigir",
        font=("Arial", 10, "bold"),
        width=15,
        bg="#4CAF50",
        fg="white",
        command=on_confirmar,
    )
    btn_confirmar.pack(side=tk.RIGHT)

    janela.bind("<Return>", lambda e: on_confirmar())
    janela.bind("<Escape>", lambda e: on_cancelar())

    janela.wait_window()

    return resultado["confirmado"], resultado["valor"]


def salvar_json_corrigido(dados, indice_config=0):
    """
    Salva os dados corrigidos no arquivo JSON.

    Args:
        dados: Lista de configurações atualizadas
        indice_config: Índice da configuração no array JSON
    """
    with open(CAMINHO_JSON, "w", encoding="utf-8") as f:
        json.dump(dados, f, indent=2, ensure_ascii=False)


def validar_estrutura_base(workbook, dados, erros):
    """
    Valida a estrutura básica: workbook e dados JSON.

    Args:
        workbook: Objeto workbook do openpyxl
        dados: Dicionário com configurações do arquivo JSON
        erros: Lista de erros (será adicionada mensagens de erro)

    Returns:
        bool: True se válido, False se inválido
    """
    if workbook is None:
        erros.append(
            "ERRO: O arquivo Excel não pôde ser aberto. Verifique se o arquivo existe e não está corrompido."
        )
        return False

    if not hasattr(workbook, "sheetnames"):
        erros.append("ERRO: O arquivo Excel está em um formato inválido ou corrompido.")
        return False

    if dados is None:
        erros.append(
            "ERRO: As configurações do sistema não foram carregadas. Entre em contato com o suporte."
        )
        return False

    if not isinstance(dados, dict):
        erros.append(
            "ERRO: As configurações do sistema estão em um formato inesperado. Entre em contato com o suporte."
        )
        return False

    return True


def validar_nome_planilha(
    workbook,
    nome_planilha,
    nome_exibicao,
    erros,
    dados,
    indice_config,
    abas_disponiveis,
):
    """
    Valida que o nome da planilha não está vazio e existe no arquivo.

    Args:
        workbook: Objeto workbook do openpyxl
        nome_planilha: Nome da planilha a validar
        nome_exibicao: Nome para exibição nas mensagens de erro
        erros: Lista de erros
        dados: Dados do JSON (para correção)
        indice_config: Índice da configuração
        abas_disponiveis: Lista de abas disponíveis no Excel

    Returns:
        tuple: (válido: bool, sheet: Worksheet ou None, corrigido: bool)
    """
    if nome_planilha is None or nome_planilha.strip() == "":
        instrucao = (
            "1. Abra o arquivo Excel\n"
            "2. Observe o nome da aba na parte inferior da tela\n"
            "3. Clique na aba que deseja usar\n"
            "4. Digite o nome exatamente como aparece"
        )
        confirmado, novo_valor = janela_corrigir_valor(
            titulo="Nome da aba não definido",
            mensagem=f"O nome da aba '{nome_exibicao}' não foi definido nas configurações.",
            instrucao=instrucao,
            valor_atual="",
            valor_default="PLANILHA ORCAMENTARIA",
        )
        if confirmado and novo_valor:
            dados[indice_config][
                (
                    "planilhaOrcamentaria"
                    if "Orçament" in nome_exibicao
                    else (
                        "planilhaFator"
                        if "Resumo" in nome_exibicao
                        else (
                            "planilhaComposicao"
                            if "Composi" in nome_exibicao
                            else "planilhaAuxiliar"
                        )
                    )
                )
            ] = novo_valor
            salvar_json_corrigido(dados, indice_config)
            return True, None, True
        return False, None, False

    if nome_planilha not in workbook.sheetnames:
        lista_abas = (
            ", ".join(workbook.sheetnames) if workbook.sheetnames else "nenhuma"
        )
        instrucao = (
            f"1. Abra o arquivo Excel\n"
            f"2. Observe as abas na parte inferior da tela\n"
            f"3. As abas disponíveis são: {lista_abas}\n"
            f"4. Digite o nome correto da aba"
        )
        confirmado, novo_valor = janela_corrigir_valor(
            titulo="Aba não encontrada",
            mensagem=f"A aba '{nome_exibicao}' não foi encontrada no arquivo Excel.",
            instrucao=instrucao,
            valor_atual=nome_planilha,
            valor_default=lista_abas.split(",")[0].strip() if "," in lista_abas else "",
        )
        if confirmado and novo_valor:
            if novo_valor in workbook.sheetnames:
                if "Orçament" in nome_exibicao:
                    dados[indice_config]["planilhaOrcamentaria"] = novo_valor
                elif "Resumo" in nome_exibicao:
                    dados[indice_config]["planilhaFator"] = novo_valor
                elif "Composi" in nome_exibicao:
                    dados[indice_config]["planilhaComposicao"] = novo_valor
                elif "Auxili" in nome_exibicao:
                    dados[indice_config]["planilhaAuxiliar"] = novo_valor
                salvar_json_corrigido(dados, indice_config)
                return True, None, True
            else:
                messagebox.showerror(
                    "Erro", f"A aba '{novo_valor}' também não existe no arquivo!"
                )
                return False, None, False
        return False, None, False

    return True, workbook[nome_planilha], False


def validar_coluna_existe(
    sheet,
    nome_coluna,
    nome_exibicao,
    erros,
    dados,
    indice_config,
    nome_planilha,
    campo_json,
):
    """
    Valida que uma coluna é válida (letra de coluna válida).

    Args:
        sheet: Worksheet do openpyxl
        nome_coluna: Letra da coluna (ex: 'A', 'B', 'AA')
        nome_exibicao: Nome para exibição nas mensagens de erro
        erros: Lista de erros
        dados: Dados do JSON (para correção)
        indice_config: Índice da configuração
        nome_planilha: Nome da aba para instrução
        campo_json: Nome do campo no JSON para correção

    Returns:
        bool: True se válido, False se inválido
    """
    if not nome_coluna or nome_coluna.strip() == "":
        instrucao = (
            f"1. Abra o arquivo Excel\n"
            f"2. Vá até a aba '{nome_planilha}'\n"
            f"3. Observe as letras no topo das colunas (A, B, C, D...)\n"
            f"4. Digite a letra da coluna que contém '{nome_exibicao}'"
        )
        confirmado, novo_valor = janela_corrigir_valor(
            titulo="Coluna não definida",
            mensagem=f"A coluna '{nome_exibicao}' não foi definida nas configurações.",
            instrucao=instrucao,
            valor_atual="",
            valor_default="A",
        )
        if confirmado and novo_valor:
            dados[indice_config][campo_json] = novo_valor.upper()
            salvar_json_corrigido(dados, indice_config)
            return True
        return False

    try:
        column_index_from_string(nome_coluna)
    except Exception:
        instrucao = (
            f"1. Abra o arquivo Excel\n"
            f"2. Vá até a aba '{nome_planilha}'\n"
            f"3. Observe as letras no topo das colunas (A, B, C, D...)\n"
            f"4. Digite a letra da coluna correta para '{nome_exibicao}'"
        )
        confirmado, novo_valor = janela_corrigir_valor(
            titulo="Coluna inválida",
            mensagem=f"A coluna '{nome_coluna}' não é uma coluna válida do Excel.",
            instrucao=instrucao,
            valor_atual=nome_coluna,
            valor_default="A",
        )
        if confirmado and novo_valor:
            dados[indice_config][campo_json] = novo_valor.upper()
            salvar_json_corrigido(dados, indice_config)
            return True
        return False

    return True


def validar_celula_bdi(
    sheet,
    coluna_fator,
    linha_fator,
    erros,
    dados,
    indice_config,
    nome_planilha="RESUMO",
):
    """
    Valida que a célula do BDI existe e tem um valor numérico.

    Args:
        sheet: Worksheet do openpyxl
        coluna_fator: Letra da coluna do BDI (ex: 'G')
        linha_fator: Número da linha do BDI (ex: 4)
        erros: Lista de erros
        dados: Dados do JSON (para correção)
        indice_config: Índice da configuração
        nome_planilha: Nome da aba para instrução

    Returns:
        tuple: (válido: bool, valor_bdi: float ou None, corrigido: bool)
    """
    corrigido = False
    linha = 0

    if not coluna_fator or not linha_fator:
        instrucao = (
            f"1. Abra o arquivo Excel\n"
            f"2. Vá até a aba '{nome_planilha}'\n"
            f"3. Procure pela linha que contém 'BDI' ou 'TAXA BDI'\n"
            f"4. Digite o número da linha (geralmente 4 ou 5)"
        )
        confirmado, novo_valor = janela_corrigir_valor(
            titulo="Localização do BDI não definida",
            mensagem="A localização do BDI (taxa de benefícios) não foi configurada corretamente.",
            instrucao=instrucao,
            valor_atual=f"Coluna: {coluna_fator}, Linha: {linha_fator}",
            valor_default="4",
        )
        if confirmado and novo_valor:
            try:
                dados[indice_config]["linhaFator"] = str(int(novo_valor))
                salvar_json_corrigido(dados, indice_config)
                linha_fator = int(novo_valor)
                corrigido = True
            except ValueError:
                messagebox.showerror(
                    "Erro", f"O valor '{novo_valor}' não é um número válido!"
                )
                return False, None, False
        else:
            return False, None, False

    try:
        linha = int(linha_fator)
        if linha <= 0:
            instrucao = (
                f"1. Abra o arquivo Excel\n"
                f"2. Vá até a aba '{nome_planilha}'\n"
                f"3. Observe o número à esquerda da linha do BDI\n"
                f"4. Digite um número maior que zero"
            )
            confirmado, novo_valor = janela_corrigir_valor(
                titulo="Linha do BDI inválida",
                mensagem=f"O número da linha do BDI ({linha_fator}) deve ser maior que zero.",
                instrucao=instrucao,
                valor_atual=str(linha_fator),
                valor_default="4",
            )
            if confirmado and novo_valor:
                try:
                    dados[indice_config]["linhaFator"] = str(int(novo_valor))
                    salvar_json_corrigido(dados, indice_config)
                    linha_fator = int(novo_valor)
                    corrigido = True
                except ValueError:
                    return False, None, False
            else:
                return False, None, False
    except (ValueError, TypeError):
        instrucao = (
            f"1. Abra o arquivo Excel\n"
            f"2. Vá até a aba '{nome_planilha}'\n"
            f"3. Observe o número à esquerda da linha do BDI\n"
            f"4. Digite apenas números (exemplo: 4)"
        )
        confirmado, novo_valor = janela_corrigir_valor(
            titulo="Linha do BDI inválida",
            mensagem=f"O valor '{linha_fator}' não é um número válido para a linha do BDI.",
            instrucao=instrucao,
            valor_atual=str(linha_fator),
            valor_default="4",
        )
        if confirmado and novo_valor:
            try:
                dados[indice_config]["linhaFator"] = str(int(novo_valor))
                salvar_json_corrigido(dados, indice_config)
                linha_fator = int(novo_valor)
                corrigido = True
            except ValueError:
                messagebox.showerror(
                    "Erro", f"O valor '{novo_valor}' não é um número válido!"
                )
                return False, None, False
        else:
            return False, None, False

    cell = sheet[f"{coluna_fator}{linha}"]
    if cell.value is None:
        print(
            f">>> [AVISO] Célula BDI (coluna {coluna_fator}, linha {linha}) está vazia. Será preenchida automaticamente."
        )
        return True, None, False

    try:
        valor_bdi = float(str(cell.value).replace(",", "."))
        return True, valor_bdi, corrigido
    except (ValueError, TypeError):
        instrucao = (
            f"1. Abra o arquivo Excel\n"
            f"2. Vá até a aba '{nome_planilha}'\n"
            f"3. Procure pela célula do BDI\n"
            f"4. Digite o valor numérico do BDI (exemplo: 28,55)"
        )
        confirmado, novo_valor = janela_corrigir_valor(
            titulo="Valor do BDI inválido",
            mensagem=f"O valor do BDI '{cell.value}' não é um número válido. O BDI deve ser um número (exemplo: 28,55).",
            instrucao=instrucao,
            valor_atual=str(cell.value),
            valor_default="28,55",
        )
        if confirmado and novo_valor:
            try:
                valor_bdi = float(str(novo_valor).replace(",", "."))
                dados[indice_config]["BDI"] = str(valor_bdi).replace(".", ",")
                salvar_json_corrigido(dados, indice_config)
                return True, valor_bdi, True
            except ValueError:
                messagebox.showerror(
                    "Erro", f"O valor '{novo_valor}' não é um número válido!"
                )
                return False, None, False
        return False, None, False


def validar_valor_existe_na_coluna(
    sheet,
    coluna,
    valor_buscado,
    nome_valor,
    nome_planilha,
    erros,
    dados,
    indice_config,
    campo_json,
):
    """
    Valida que um valor existe em uma coluna específica.

    Args:
        sheet: Worksheet do openpyxl
        coluna: Letra da coluna para buscar
        valor_buscado: Valor a buscar
        nome_valor: Nome do valor para exibição
        nome_planilha: Nome da planilha para mensagens de erro
        erros: Lista de erros
        dados: Dados do JSON (para correção)
        indice_config: Índice da configuração
        campo_json: Nome do campo no JSON para correção

    Returns:
        bool: True se encontrado, False se não encontrado
    """
    if not valor_buscado or valor_buscado.strip() == "":
        return True

    linha_encontrada = buscar_palavra(sheet, coluna, valor_buscado)

    if linha_encontrada == -1:
        instrucao = (
            f"1. Abra o arquivo Excel\n"
            f"2. Vá até a aba '{nome_planilha}'\n"
            f"3. Procure na coluna '{coluna}' pelo texto relacionado a '{nome_valor}'\n"
            f"4. Digite o texto exatamente como aparece na célula"
        )
        while True:
            confirmado, novo_valor = janela_corrigir_valor(
                titulo="Texto não encontrado",
                mensagem=f"O texto '{valor_buscado}' não foi encontrado na coluna '{coluna}' da aba '{nome_planilha}'.",
                instrucao=instrucao,
                valor_atual=valor_buscado,
                valor_default=valor_buscado,
            )
            if not confirmado:
                return False
            if not novo_valor or novo_valor.strip() == "":
                messagebox.showerror(
                    "Erro", "O valor não pode ser vazio. Digite um valor válido."
                )
                continue
            linha_validada = buscar_palavra(sheet, coluna, novo_valor)
            if linha_validada == -1:
                messagebox.showerror(
                    "Erro",
                    f"O texto '{novo_valor}' não foi encontrado na coluna '{coluna}'.\n"
                    f"Verifique se o texto está correto e tente novamente.",
                )
                valor_buscado = novo_valor
                continue
            dados[indice_config][campo_json] = novo_valor
            salvar_json_corrigido(dados, indice_config)
            return True

    return True


def validar_valor_na_coluna(
    sheet,
    coluna_atual,
    valor_atual,
    nome_valor_exibicao,
    nome_planilha,
    dados,
    indice_config,
    campo_json_coluna,
    campo_json_valor,
):
    while True:
        instrucao_coluna = (
            f"1. Abra o arquivo Excel\n"
            f"2. Vá até a aba '{nome_planilha}'\n"
            f"3. Observe as letras no topo das colunas (A, B, C, D...)\n"
            f"4. Digite a letra da coluna que contém o {nome_valor_exibicao}"
        )
        confirmado_coluna, nova_coluna = janela_corrigir_valor(
            titulo=f"Coluna do {nome_valor_exibicao}",
            mensagem=f"O texto do {nome_valor_exibicao} não foi encontrado na coluna '{coluna_atual}'.\n"
            f"Verifique se a COLUNA está correta!",
            instrucao=instrucao_coluna,
            valor_atual=coluna_atual,
            valor_default=coluna_atual,
        )
        if not confirmado_coluna:
            return False

        try:
            column_index_from_string(nova_coluna.upper())
        except Exception:
            messagebox.showerror("Erro", f"Coluna '{nova_coluna}' inválida!")
            continue

        instrucao_valor = (
            f"1. Abra o arquivo Excel\n"
            f"2. Vá até a aba '{nome_planilha}'\n"
            f"3. Vá até a coluna '{nova_coluna.upper()}' e procure pelo {nome_valor_exibicao}\n"
            f"4. Digite o texto EXATAMENTE como aparece na célula"
        )
        confirmado_valor, novo_valor = janela_corrigir_valor(
            titulo="Texto não encontrado",
            mensagem=f"O texto '{valor_atual}' não foi encontrado na coluna '{nova_coluna}'.\n"
            f"Digite o texto correto que aparece no Excel.",
            instrucao=instrucao_valor,
            valor_atual=valor_atual,
            valor_default=valor_atual,
        )
        if not confirmado_valor:
            return False

        linha_encontrada = buscar_palavra(sheet, nova_coluna.upper(), novo_valor)

        if linha_encontrada != -1:
            dados[indice_config][campo_json_coluna] = nova_coluna.upper()
            dados[indice_config][campo_json_valor] = novo_valor
            salvar_json_corrigido(dados, indice_config)
            return True

        messagebox.showerror(
            "Erro",
            f"O texto '{novo_valor}' não foi encontrado na coluna '{nova_coluna}'.\n"
            f"Tente novamente com valores corretos.",
        )


# ============================================
# FUNÇÕES DE VALIDAÇÃO POR PLANILHA
# ============================================


def validar_planilha_orcamentaria(workbook, dados, erros, indice_config=0):
    """
    Valida a planilha orçamentária com correção interativa.

    Args:
        workbook: Objeto workbook do openpyxl
        dados: Dicionário com configurações do arquivo JSON
        erros: Lista de erros
        indice_config: Índice da configuração no JSON

    Returns:
        tuple: (válido: bool, sheet: Worksheet ou None, linha_cabecalhos: int ou None)
    """
    nome_planilha = dados[indice_config].get(
        "planilhaOrcamentaria", "PLANILHA ORCAMENTARIA"
    )

    valido, sheet, corrigido = validar_nome_planilha(
        workbook,
        nome_planilha,
        "Orçamentária",
        erros,
        dados,
        indice_config,
        workbook.sheetnames if hasattr(workbook, "sheetnames") else [],
    )
    if not valido:
        return False, None, None

    if sheet and sheet.max_row < 2:
        erros.append(
            f"ERRO: A aba '{nome_planilha}' está vazia ou não tem dados suficientes."
        )
        return False, sheet, None

    # Carregar valores atuais do JSON (pode ter sido corrigido em iterations anteriores)
    dados = json.load(open(CAMINHO_JSON, "r", encoding="utf-8"))

    coluna_inicial = dados[indice_config].get("colunaInicial", "A")
    valor_inicial = dados[indice_config].get("valorInicial", "ITEM")
    valor_final = dados[indice_config].get("valorFinal", "VALOR BDI TOTAL")

    if sheet:
        if not validar_coluna_existe(
            sheet,
            coluna_inicial,
            "Inicial",
            erros,
            dados,
            indice_config,
            nome_planilha,
            "colunaInicial",
        ):
            return False, sheet, None

        coluna_final = dados[indice_config].get("colunaFinal", "F")
        if not validar_coluna_existe(
            sheet,
            coluna_final,
            "Coluna Final",
            erros,
            dados,
            indice_config,
            nome_planilha,
            "colunaFinal",
        ):
            return False, sheet, None

        # Tentar encontrar o valor inicial na coluna
        linha_cabecalhos = buscar_palavra(sheet, coluna_inicial, valor_inicial)

        if linha_cabecalhos == -1:
            # Primeiro, perguntar se a COLUNA está correta
            instrucao_coluna = (
                f"1. Abra o arquivo Excel\n"
                f"2. Vá até a aba '{nome_planilha}'\n"
                f"3. Observe as letras no topo das colunas (A, B, C, D...)\n"
                f"4. Digite a letra da coluna que contém '{valor_inicial}' ou 'ITEM'"
            )
            confirmado_coluna, nova_coluna = janela_corrigir_valor(
                titulo="Coluna do valor inicial",
                mensagem=f"O valor '{valor_inicial}' não foi encontrado na coluna '{coluna_inicial}'.\n"
                f"Verifique se a COLUNA está correta!",
                instrucao=instrucao_coluna,
                valor_atual=coluna_inicial,
                valor_default="A",
            )

            if confirmado_coluna and nova_coluna:
                # Validar se a nova coluna é válida
                try:
                    column_index_from_string(nova_coluna.upper())
                    dados[indice_config]["colunaInicial"] = nova_coluna.upper()
                    salvar_json_corrigido(dados, indice_config)
                    coluna_inicial = nova_coluna.upper()
                except Exception:
                    messagebox.showerror("Erro", f"Coluna '{nova_coluna}' inválida!")
                    return False, sheet, None

            # Agora pedir o valor correto
            instrucao_valor = (
                f"1. Abra o arquivo Excel\n"
                f"2. Vá até a aba '{nome_planilha}'\n"
                f"3. Vá até a coluna '{coluna_inicial}' e procure pelo texto 'ITEM' ou cabeçalho inicial\n"
                f"4. Digite o texto EXATAMENTE como aparece na célula"
            )
            confirmado_valor, novo_valor = janela_corrigir_valor(
                titulo="Texto inicial não encontrado",
                mensagem=f"O texto '{valor_inicial}' não foi encontrado na coluna '{coluna_inicial}'.\n"
                f"Digite o texto correto que aparece no Excel.",
                instrucao=instrucao_valor,
                valor_atual=valor_inicial,
                valor_default="ITEM",
            )

            if confirmado_valor and novo_valor:
                dados[indice_config]["valorInicial"] = novo_valor
                salvar_json_corrigido(dados, indice_config)

                # Tentar encontrar novamente com o novo valor
                linha_cabecalhos = buscar_palavra(sheet, coluna_inicial, novo_valor)
                if linha_cabecalhos == -1:
                    erros.append(
                        f"ERRO: O texto '{novo_valor}' ainda não foi encontrado na coluna '{coluna_inicial}'."
                    )
                    return False, sheet, None
            else:
                return False, sheet, None

        valores_linha = []
        for cell in sheet[linha_cabecalhos + 1]:
            if cell.value is not None:
                valores_linha.append(str(cell.value).strip().upper())
            else:
                valores_linha.append("")

        cabecalhos_esperados = ["ITEM", "CÓDIGO", "DESCRIÇÃO", "UND", "QUANTIDADE"]

        cabecalhos_faltantes = []
        for cabecalho in cabecalhos_esperados:
            encontrado = False
            for valor in valores_linha:
                if cabecalho.upper() in valor or valor in cabecalho.upper():
                    encontrado = True
                    break
            if not encontrado:
                cabecalhos_faltantes.append(cabecalho)

        if len(cabecalhos_faltantes) > 0:
            erros.append(
                f"ERRO: Algumas colunas obrigatórias não foram encontradas na linha {linha_cabecalhos + 1} da aba '{nome_planilha}'.\n"
                f"Colunas esperadas: {', '.join(cabecalhos_esperados)}\n"
                f"Colunas encontradas: {', '.join(valores_linha)}\n"
                f"Faltando: {', '.join(cabecalhos_faltantes)}"
            )
            return False, sheet, linha_cabecalhos

        if valor_final:
            if not validar_valor_na_coluna(
                sheet,
                coluna_final,
                valor_final,
                "VALOR TOTAL",
                nome_planilha,
                dados,
                indice_config,
                "colunaFinal",
                "valorFinal",
            ):
                return False, sheet, linha_cabecalhos

        return True, sheet, linha_cabecalhos

    return True, sheet, None


def validar_planilha_resumo(workbook, dados, erros, indice_config=0):
    """
    Valida a planilha RESUMO (FATOR) com correção interativa.

    Args:
        workbook: Objeto workbook do openpyxl
        dados: Dicionário com configurações do arquivo JSON
        erros: Lista de erros
        indice_config: Índice da configuração no JSON

    Returns:
        tuple: (válido: bool, sheet: Worksheet ou None)
    """
    nome_planilha = dados[indice_config].get("planilhaFator", "RESUMO")

    valido, sheet, _ = validar_nome_planilha(
        workbook,
        nome_planilha,
        "Resumo",
        erros,
        dados,
        indice_config,
        workbook.sheetnames if hasattr(workbook, "sheetnames") else [],
    )
    if not valido:
        return False, None

    if sheet and sheet.max_row < 2:
        erros.append(
            f"ERRO: A aba '{nome_planilha}' está vazia ou não tem dados suficientes."
        )
        return False, sheet

    coluna_fator = dados[indice_config].get("colunaFator", "G")
    linha_fator = dados[indice_config].get("linhaFator", "4")

    if sheet:
        if not validar_coluna_existe(
            sheet,
            coluna_fator,
            "Fator",
            erros,
            dados,
            indice_config,
            nome_planilha,
            "colunaFator",
        ):
            return False, sheet

        bdi_valido, _, _ = validar_celula_bdi(
            sheet, coluna_fator, linha_fator, erros, dados, indice_config, nome_planilha
        )
        if not bdi_valido:
            return False, sheet

        valor_total_resumo = dados[indice_config].get(
            "valorTotalResumo", "VALOR TOTAL RESUMO:"
        )
        coluna_total_resumo = dados[indice_config].get("colunaTotalResumo", "C")

        if valor_total_resumo:
            if not validar_valor_na_coluna(
                sheet,
                coluna_total_resumo,
                valor_total_resumo,
                "VALOR TOTAL DO RESUMO",
                nome_planilha,
                dados,
                indice_config,
                "colunaTotalResumo",
                "valorTotalResumo",
            ):
                return False, sheet

    return True, sheet


def validar_planilha_composicoes(workbook, dados, erros, indice_config=0):
    """
    Valida a planilha COMPOSIÇÕES com correção interativa.

    Args:
        workbook: Objeto workbook do openpyxl
        dados: Dicionário com configurações do arquivo JSON
        erros: Lista de erros
        indice_config: Índice da configuração no JSON

    Returns:
        tuple: (válido: bool, sheet: Worksheet ou None)
    """
    nome_planilha = dados[indice_config].get("planilhaComposicao", "COMPOSICOES")

    valido, sheet, _ = validar_nome_planilha(
        workbook,
        nome_planilha,
        "Composições",
        erros,
        dados,
        indice_config,
        workbook.sheetnames if hasattr(workbook, "sheetnames") else [],
    )
    if not valido:
        return False, None

    if sheet and sheet.max_row < 2:
        erros.append(
            f"ERRO: A aba '{nome_planilha}' está vazia ou não tem dados suficientes."
        )
        return False, sheet

    if sheet:
        colunas_comp = [
            ("composicaoDescricao", "Descrição", "A"),
            ("colunaItemDescricaoComposicao", "Código do Item", "B"),
            ("composicaoCoeficiente", "Coeficiente", "E"),
            ("composicaoPrecoUnitario", "Preço Unitário", "F"),
            ("composicaoCoeficienteCopiar", "Coeficiente (copiar)", "L"),
            ("composicaoPrecoUnitarioCopiar", "Preço Unitário (copiar)", "M"),
            ("colunaTotaisComposicao", "Coluna de Totais", "E"),
        ]

        for campo_json, nome_col, default_col in colunas_comp:
            col = dados[indice_config].get(campo_json, default_col)
            if not validar_coluna_existe(
                sheet,
                col,
                nome_col,
                erros,
                dados,
                indice_config,
                nome_planilha,
                campo_json,
            ):
                return False, sheet

        col_totais = get_coluna_totais_comp(dados[indice_config])
        col_valor_totais = get_valor_totais_comp(dados[indice_config])

        if not validar_coluna_existe(
            sheet,
            col_totais,
            "Coluna de Totais",
            erros,
            dados,
            indice_config,
            nome_planilha,
            "colunaTotaisComposicao",
        ):
            return False, sheet

        if not validar_coluna_existe(
            sheet,
            col_valor_totais,
            "Coluna de Valores Totais",
            erros,
            dados,
            indice_config,
            nome_planilha,
            "valorTotaisComposicao",
        ):
            return False, sheet

        valores_a_verificar = {
            "total_com_bdi": (
                get_valor_com_bdi_string(dados[indice_config]),
                "valorComBdi",
            ),
            "valor_bdi": (get_valor_bdi_comp(dados[indice_config]), "valorBdi"),
            "valor_string": (get_valor_string(dados[indice_config]), "valor"),
        }

        for nome_valor, (valor_buscado, campo_json) in valores_a_verificar.items():
            if not validar_valor_existe_na_coluna(
                sheet,
                col_totais,
                valor_buscado,
                nome_valor,
                nome_planilha,
                erros,
                dados,
                indice_config,
                campo_json,
            ):
                return False, sheet

    return True, sheet


def validar_planilha_composicoes_auxiliares(workbook, dados, erros, indice_config=0):
    """
    Valida a planilha COMPOSIÇÕES AUXILIARES com correção interativa.

    Args:
        workbook: Objeto workbook do openpyxl
        dados: Dicionário com configurações do arquivo JSON
        erros: Lista de erros
        indice_config: Índice da configuração no JSON

    Returns:
        tuple: (válido: bool, sheet: Worksheet ou None)
    """
    nome_planilha = dados[indice_config].get(
        "planilhaAuxiliar", "COMPOSICOES AUXILIARES"
    )

    valido, sheet, _ = validar_nome_planilha(
        workbook,
        nome_planilha,
        "Auxiliares",
        erros,
        dados,
        indice_config,
        workbook.sheetnames if hasattr(workbook, "sheetnames") else [],
    )
    if not valido:
        return False, None

    if sheet and sheet.max_row < 2:
        erros.append(
            f"ERRO: A aba '{nome_planilha}' está vazia ou não tem dados suficientes."
        )
        return False, sheet

    if sheet:
        colunas_aux = [
            ("auxiliarDescricao", "Descrição", "A"),
            ("auxiliarCoeficiente", "Coeficiente", "E"),
            ("auxiliarPrecoUnitario", "Preço Unitário", "F"),
            ("auxiliarCoeficienteCopiar", "Coeficiente (copiar)", "L"),
            ("auxiliarPrecoUnitarioCopiar", "Preço Unitário (copiar)", "M"),
            ("colunaTotaisAuxiliar", "Coluna de Totais", "E"),
        ]

        for campo_json, nome_col, default_col in colunas_aux:
            col = dados[indice_config].get(campo_json, default_col)
            if not validar_coluna_existe(
                sheet,
                col,
                nome_col,
                erros,
                dados,
                indice_config,
                nome_planilha,
                campo_json,
            ):
                return False, sheet

        col_totais = get_coluna_totais_aux(dados[indice_config])
        col_valor_totais_aux = get_valor_totais_aux(dados[indice_config])

        if not validar_coluna_existe(
            sheet,
            col_totais,
            "Coluna de Totais",
            erros,
            dados,
            indice_config,
            nome_planilha,
            "colunaTotaisAuxiliar",
        ):
            return False, sheet

        if not validar_coluna_existe(
            sheet,
            col_valor_totais_aux,
            "Coluna de Valores Totais",
            erros,
            dados,
            indice_config,
            nome_planilha,
            "valorTotaisAuxiliar",
        ):
            return False, sheet

        valores_a_verificar = {
            "total_com_bdi": (
                get_valor_com_bdi_string(dados[indice_config]),
                "valorComBdi",
            ),
            "valor_bdi": (get_valor_bdi_comp(dados[indice_config]), "valorBdi"),
            "valor_string": (get_valor_string(dados[indice_config]), "valor"),
        }

        for nome_valor, (valor_buscado, campo_json) in valores_a_verificar.items():
            if not validar_valor_existe_na_coluna(
                sheet,
                col_totais,
                valor_buscado,
                nome_valor,
                nome_planilha,
                erros,
                dados,
                indice_config,
                campo_json,
            ):
                return False, sheet

    return True, sheet


# ============================================
# FUNÇÃO PRINCIPAL DE VALIDAÇÃO
# ============================================


def validar_arquivo_excel(filepath, dados):
    """
    Valida a estrutura completa do arquivo Excel antes do processamento.

    Se encontrar erros que podem ser corrigidos, exibe janela para correção
    e salva automaticamente no arquivo JSON. Recarrega o workbook e os dados
    após cada correção para garantir que as mudanças sejam refletidas.

    Args:
        filepath: Caminho do arquivo Excel
        dados: Lista de configurações do arquivo JSON

    Returns:
        tuple: (True, workbook, dados_atualizados) se válido,
               ou (False, None, None) se inválido/cancelado
    """
    if not isinstance(dados, list):
        dados = [dados]

    indice_config = 0

    print("=" * 60)
    print(">>> INICIANDO VALIDAÇÃO DO ARQUIVO EXCEL")
    print("=" * 60)

    while True:
        erros = []
        workbook = openpyxl.load_workbook(filepath)

        dados_atualizados = json.load(open(CAMINHO_JSON, "r", encoding="utf-8"))
        json_antes = json.dumps(dados_atualizados, sort_keys=True)

        print("\n>>> [FASE 1] Validando estrutura base...")
        if not validar_estrutura_base(
            workbook, dados_atualizados[indice_config], erros
        ):
            mensagem = "ERROS NA VALIDAÇÃO:\n" + "\n".join(erros)
            return False, None, None

        json_depois = json.dumps(
            json.load(open(CAMINHO_JSON, "r", encoding="utf-8")), sort_keys=True
        )
        if json_antes != json_depois:
            print(">>> [INFO] Configurações atualizadas. Recarregando...")
            workbook.close()
            continue
        json_antes = json_depois

        print(">>> [OK] Estrutura base válida")

        print("\n>>> [FASE 2] Validando planilha orçamentária...")
        valido, sheet_orcamentaria, linha_cabecalhos = validar_planilha_orcamentaria(
            workbook, dados_atualizados, erros, indice_config
        )

        if not valido:
            print(
                ">>> [ERRO] Validação da planilha orçamentária falhou ou foi cancelada."
            )
            print(">>> Encerrando validação.")
            workbook.close()
            return False, None, None

        json_depois = json.dumps(
            json.load(open(CAMINHO_JSON, "r", encoding="utf-8")), sort_keys=True
        )
        if json_antes != json_depois:
            print(">>> [INFO] Nome da planilha orçamentária corrigido. Recarregando...")
            workbook.close()
            continue
        json_antes = json_depois

        print(
            f">>> [OK] Planilha orçamentária válida (cabeçalhos na linha {linha_cabecalhos + 1 if linha_cabecalhos else '?'})"
        )

        print("\n>>> [FASE 3] Validando planilha RESUMO...")
        valido, sheet_resumo = validar_planilha_resumo(
            workbook, dados_atualizados, erros, indice_config
        )

        if not valido:
            print(">>> [ERRO] Validação da planilha RESUMO falhou ou foi cancelada.")
            print(">>> Encerrando validação.")
            workbook.close()
            return False, None, None

        json_depois = json.dumps(
            json.load(open(CAMINHO_JSON, "r", encoding="utf-8")), sort_keys=True
        )
        if json_antes != json_depois:
            print(">>> [INFO] Nome da planilha RESUMO corrigido. Recarregando...")
            workbook.close()
            continue
        json_antes = json_depois

        print("\n>>> [FASE 4] Validando planilha COMPOSIÇÕES...")
        valido, sheet_composicao = validar_planilha_composicoes(
            workbook, dados_atualizados, erros, indice_config
        )

        if not valido:
            print(
                ">>> [ERRO] Validação da planilha COMPOSIÇÕES falhou ou foi cancelada."
            )
            print(">>> Encerrando validação.")
            workbook.close()
            return False, None, None

        json_depois = json.dumps(
            json.load(open(CAMINHO_JSON, "r", encoding="utf-8")), sort_keys=True
        )
        if json_antes != json_depois:
            print(">>> [INFO] Nome da planilha COMPOSIÇÕES corrigido. Recarregando...")
            workbook.close()
            continue
        json_antes = json_depois

        print("\n>>> [FASE 5] Validando planilha COMPOSIÇÕES AUXILIARES...")
        valido, sheet_auxiliar = validar_planilha_composicoes_auxiliares(
            workbook, dados_atualizados, erros, indice_config
        )

        if not valido:
            print(
                ">>> [ERRO] Validação da planilha AUXILIARES falhou ou foi cancelada."
            )
            print(">>> Encerrando validação.")
            workbook.close()
            return False, None, None

        json_depois = json.dumps(
            json.load(open(CAMINHO_JSON, "r", encoding="utf-8")), sort_keys=True
        )
        if json_antes != json_depois:
            print(">>> [INFO] Nome da planilha AUXILIARES corrigido. Recarregando...")
            workbook.close()
            continue

        print("\n" + "=" * 60)
        print(">>> [OK] VALIDAÇÃO CONCLUÍDA COM SUCESSO!")
        print("=" * 60)
        return True, workbook, dados_atualizados
