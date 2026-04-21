from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.cell.cell import MergedCell


def verificar_e_adicionar_fator(workbook, dados):
    """
    Verifica e adiciona fórmulas de fator para itens com 'adicionarFator': 'Sim'.

    Lógica:
    - Se fatorCoeficiente: "Sim" → usar coluna E (composicaoCoeficiente ou auxiliarCoeficiente)
    - Se fatorCoeficiente: "Não" → usar coluna F (composicaoPrecoUnitario ou auxiliarPrecoUnitario)

    Valida em ambas as planilhas: composições e composições auxiliares.
    """
    print(">>> Verificando e adicionando fatores dos itens...")

    # O dados pode ser uma lista [{}] ou um dicionário {}
    dados_itens = dados
    if isinstance(dados, list) and len(dados) > 0:
        dados_itens = dados[0]

    # Obter nomes das planilhas do JSON
    planilha_comp = dados_itens.get("planilhaComposicao", "COMPOSICOES")
    planilha_aux = dados_itens.get("planilhaAuxiliar", "COMPOSICOES AUXILIARES")

    # Obter colunas do JSON
    col_coef_comp = dados_itens.get("composicaoCoeficiente", "E")
    col_preco_comp = dados_itens.get("composicaoPrecoUnitario", "F")
    col_coef_comp_antigo = dados_itens.get("composicaoCoeficienteCopiar", "L")
    col_preco_comp_antigo = dados_itens.get("composicaoPrecoUnitarioCopiar", "M")
    col_desc_comp = dados_itens.get("composicaoDescricao", "A")

    col_coef_aux = dados_itens.get("auxiliarCoeficiente", "E")
    col_preco_aux = dados_itens.get("auxiliarPrecoUnitario", "F")
    col_coef_aux_antigo = dados_itens.get("auxiliarCoeficienteCopiar", "L")
    col_preco_aux_antigo = dados_itens.get("auxiliarPrecoUnitarioCopiar", "M")
    col_desc_aux = dados_itens.get("auxiliarDescricao", "A")

    # Obter valores de linha a pular (VALOR, VALOR BDI, VALOR COM BDI)
    valor_label = dados_itens.get("valor", "VALOR:")
    valor_bdi_label = dados_itens.get("valorBdi", "VALOR BDI")
    valor_com_bdi_label = dados_itens.get("valorComBdi", "VALOR COM BDI")

    # Construir lista de valores a pular (texto que indica linhas de totals/valores)
    valores_pular = [
        valor_label.upper(),
        valor_bdi_label.upper(),
        valor_com_bdi_label.upper(),
    ]

    # Construir lista de itens que precisam de fator (MANTER INDIVIDUAL)
    # Cada item é processado separadamente com seus próprios filtros
    # IMPORTANTE: Não incluir itens com buscarAuxiliar: "Sim" pois eles devem receber
    # fórmula de referência (buscar_auxiliar), não fórmula de fator
    itens_fator = []

    for key, item in dados_itens.items():
        if key.startswith("item") and isinstance(item, dict):
            # Pular itens que têm buscarAuxiliar: "Sim" - eles recebem fórmula de referência
            if item.get("buscarAuxiliar") == "Sim":
                continue
            if item.get("adicionarFator") == "Sim":
                itens_fator.append(
                    {
                        "nome": item.get("nome", ""),
                        "total": item.get("total", ""),
                        "fatorCoeficiente": item.get("fatorCoeficiente") == "Sim",
                        "iniciaPor": item.get("iniciaPor", ""),
                        "naoIniciaPor": item.get("naoIniciaPor", ""),
                    }
                )

    print(f">> Itens com adicionarFator: Sim - {len(itens_fator)}")
    for item in itens_fator:
        print(
            f"   {item['nome']}: fatorCoeficiente={item['fatorCoeficiente']}, "
            f"iniciaPor='{item['iniciaPor']}', naoIniciaPor='{item['naoIniciaPor']}'"
        )

    # Labels para pular (itens com buscarAuxiliar: "Não")
    labels_pular = []
    totals_pular = []
    for key, item in dados_itens.items():
        if key.startswith("item") and isinstance(item, dict):
            if item.get("buscarAuxiliar") == "Não":
                nome = item.get("nome", "")
                if nome:
                    labels_pular.append(nome.upper())
                total_str = item.get("total", "")
                if total_str:
                    totals_pular.append(total_str.upper())

    print(f">> Labels para pular (buscarAuxiliar: Não): {labels_pular}")
    print(f">> Totals para pular: {totals_pular}")

    total_adicionados_comp = 0
    total_adicionados_aux = 0

    # Verificar composições
    print(f"\n>> Verificando planilha: {planilha_comp}")
    # ✅ NA PLANILHA PRINCIPAL COMPOSIÇÕES: PROCESSAR TODOS OS ITENS, INCLUSIVE OS COM buscarAuxiliar: Não
    adicionados = verificar_e_adicionar_planilha(
        workbook[planilha_comp],
        column_index_from_string(col_desc_comp),
        column_index_from_string(col_coef_comp),
        column_index_from_string(col_preco_comp),
        column_index_from_string(col_coef_comp_antigo),
        column_index_from_string(col_preco_comp_antigo),
        itens_fator,
        [],  # NAO PULAR NADA NA PLANILHA PRINCIPAL
        totals_pular,
        [],  # NAO PULAR VALORES NA PLANILHA PRINCIPAL
    )
    total_adicionados_comp = adicionados
    print(f">> Fórmulas adicionadas em Composições: {adicionados}")

    # Verificar composições auxiliares
    print(f"\n>> Verificando planilha: {planilha_aux}")
    # ✅ PROCESSAR TODOS OS ITENS com adicionarFator: "Sim" NA PLANILHA AUXILIAR
    # O filtro buscarAuxiliar: "Não" é irrelevante para a aplicação de fórmulas de fator
    itens_fator_aux = itens_fator
    adicionados = verificar_e_adicionar_planilha(
        workbook[planilha_aux],
        column_index_from_string(col_desc_aux),
        column_index_from_string(col_coef_aux),
        column_index_from_string(col_preco_aux),
        column_index_from_string(col_coef_aux_antigo),
        column_index_from_string(col_preco_aux_antigo),
        itens_fator_aux,
        [],
        totals_pular,
        valores_pular,
    )
    total_adicionados_aux = adicionados
    print(f">> Fórmulas adicionadas em Composições Auxiliares: {adicionados}")

    # Resumo
    total = total_adicionados_comp + total_adicionados_aux
    print(f"\n>>> Total de fórmulas de fator adicionadas: {total}")

    return total


def encontrar_todas_secoes(
    sheet, col_desc_idx, nome_upper, total_upper, inicia_por="", nao_inicia_por=""
):
    """
    Encontra TODAS as seções com o nome dado na planilha.
    Retorna lista de tuplas (inicio, fim) para cada seção.
    ONLY retorna seções onde o título começa com inicia_por (se fornecido)
    e NÃO começa com nao_inicia_por (se fornecido).
    """
    secoes = []
    colunas_busca = list(range(col_desc_idx, col_desc_idx + 10))
    max_row = sheet.max_row

    # Primeiro, encontrar todos os inícios que passam nos filtros
    inicios = []
    for i in range(1, max_row + 1):
        for col_busca in colunas_busca:
            cell = sheet.cell(row=i, column=col_busca).value
            if cell and isinstance(cell, MergedCell):
                continue
            cell_str = str(cell).strip()
            cell_upper = cell_str.upper()
            if cell and nome_upper in cell_upper:
                # Verificar que não é um TOTAL
                if "TOTAL" not in cell_upper:
                    # Se inicia_por fornecido, verificar que o título começa com ele
                    # Se inicia_por vazio mas nome_upper foi fornecido, usar startswith(nome_upper)
                    if inicia_por:
                        if not cell_str.startswith(inicia_por):
                            continue
                    else:
                        # Quando inicia_por é vazio, usar startswith EXATO para evitar substring matches
                        # A seção deve começar exatamente com nome_upper (possivelmente seguido de ":")
                        if not (
                            cell_upper.startswith(nome_upper)
                            and (
                                len(cell_upper) == len(nome_upper)
                                or cell_upper[len(nome_upper)] in [" ", ":"]
                            )
                        ):
                            continue
                    # Se nao_inicia_por fornecido, verificar que o título NÃO começa com ele
                    if nao_inicia_por and cell_upper.startswith(nao_inicia_por.upper()):
                        continue
                    inicios.append((i, col_busca))
                    break

    # Para cada início, encontrar o fim mais próximo
    # ✅ CORRIGIDO: Usar endswith em vez de 'in' para evitar confusão entre seções com nomes similares
    # Ex: "Mão de Obra" não deve ser cortado pelo TOTAL de "Mão de Obra com Encargos Complementares"
    for inicio, col_inicio in inicios:
        fim = -1
        for j in range(inicio + 1, max_row + 1):
            for col_busca in colunas_busca:
                cell = sheet.cell(row=j, column=col_busca).value
                cell_upper = str(cell).upper() if cell else ""
                if cell and cell_upper.endswith(total_upper):
                    fim = j
                    break
            if fim != -1:
                break
        if fim != -1:
            secoes.append((inicio, fim))

    return secoes


def verificar_e_adicionar_planilha(
    sheet,
    col_desc_idx,
    col_coef_idx,
    col_preco_idx,
    col_coef_antigo_idx,
    col_preco_antigo_idx,
    itens_fator,
    labels_pular,
    totals_pular,
    valores_pular,
):
    """
    Verifica uma planilha específica e adiciona fórmulas de fator onde necessário.
    Processa cada item em TODAS as seções correspondentes.
    """
    linhas_adicionadas = []
    max_row = sheet.max_row

    # Encontrar TODAS as seções para cada nome único
    # NÃO usar mais nomes_processados para permitir múltiplos itens com mesmo nome
    secoes_encontradas = (
        {}
    )  # cache_key (nome_upper, iniciaPor, naoIniciaPor) -> [(inicio, fim), ...]

    # Primeiro pass: encontrar todas as seções únicas (SEM FILTRO de iniciaPor/naoIniciaPor no título)
    # O filtro de código será aplicado quando processar as linhas dentro das seções
    nomes_processados = set()
    for item_info in itens_fator:
        nome = item_info["nome"]
        total_str = item_info["total"]
        nome_upper = nome.upper()
        total_upper = total_str.upper() if total_str else ""

        # Criar chave única só pelo nome (sem iniciaPor/naoIniciaPor)
        cache_key = nome_upper

        # Se já buscamos essa seção, não buscar de novo
        if cache_key in nomes_processados:
            continue

        nomes_processados.add(cache_key)

        # Buscar todas as seções com esse nome (SEM filtro iniciaPor/naoIniciaPor)
        # O filtro de código será aplicado quando processar as linhas
        secoes = encontrar_todas_secoes(
            sheet, col_desc_idx, nome_upper, total_upper, "", ""
        )
        if secoes:
            secoes_encontradas[cache_key] = secoes
            print(f"   Encontradas {len(secoes)} seções de '{nome_upper}'")

    # Processar cada item em TODAS as seções correspondentes
    for item_info in itens_fator:
        nome = item_info["nome"]
        total_str = item_info["total"]
        fator_coef = item_info["fatorCoeficiente"]
        inicia_por = item_info["iniciaPor"]
        nao_inicia_por = item_info["naoIniciaPor"]

        nome_upper = nome.upper()
        # ✅ CORRIGIDO: Usar cache_key simples (só nome_upper) para acessar as seções
        # que foram armazenadas com a mesma chave simples na linha 238
        cache_key = nome_upper

        # Obter lista de seções para este nome
        if cache_key not in secoes_encontradas:
            print(
                f"   !! Não encontrou nenhuma seção para: {nome_upper} (iniciaPor='{inicia_por}', naoIniciaPor='{nao_inicia_por}')"
            )
            continue

        secoes = secoes_encontradas[cache_key]
        count_processadas = 0

        # Processar TODAS as seções que correspondem ao filtro
        for idx, (inicio, fim) in enumerate(secoes):
            count_processadas += 1
            print(
                f"   Processando [{count_processadas}]: {nome_upper} (L{inicio} a L{fim}), "
                f"iniciaPor='{inicia_por}', naoIniciaPor='{nao_inicia_por}'"
            )

            # Processar linhas entre início e fim
            for y in range(inicio + 1, fim):
                cell_desc = sheet.cell(row=y, column=col_desc_idx).value
                if isinstance(cell_desc, MergedCell):
                    continue

                desc = str(cell_desc) if cell_desc else ""

                # Pular linhas de títulos de seção (que contêm palavras como FONTE, UNID, COEFICIENTE)
                desc_upper = desc.upper()
                if "COEFICIENTE" in desc_upper or "PREÇO UNITÁRIO" in desc_upper:
                    continue

                # Verificar se é uma linha de título de seção (contém FONTE ou UNID ou TOTAL)
                cell_fonte = sheet.cell(row=y, column=col_desc_idx + 2).value
                cell_unid = sheet.cell(row=y, column=col_desc_idx + 3).value
                if cell_fonte and "FONTE" in str(cell_fonte).upper():
                    continue
                if cell_unid and "UNID" in str(cell_unid).upper():
                    continue

                # NÃO pular linhas só porque a seção tem um nome em labels_pular
                # O labels_pular serve para outros contextos, não para pular a seção inteira
                # Apenas pular linhas que são especificamente o label (título da seção)
                # e linhas de TOTAL
                # ✅ CORREÇÃO: labels_pular NÃO deve pular linhas INTERNAS da seção
                # Removida verificação que pulava TODAS as linhas de itens com buscarAuxiliar: "Não"

                # Apenas pular linhas de TOTAL (verificar coluna A E coluna E onde os totais aparecem)
                if any(x in desc_upper for x in totals_pular):
                    continue

                # Pular linhas que contêm TOTAL, VALOR, VALOR BDI, VALOR COM BDI em qualquer checking
                if any(
                    x in desc_upper
                    for x in ["TOTAL", "VALOR", "VALOR BDI", "VALOR COM BDI"]
                ):
                    continue

                # Verificar também a coluna E (col_coef_idx) onde os totais aparecem
                cell_coef_check = sheet.cell(row=y, column=col_coef_idx).value
                if cell_coef_check:
                    coef_upper = str(cell_coef_check).upper()
                    if any(x in coef_upper for x in totals_pular):
                        continue
                    # Pular linhas que contêm TOTAL, VALOR, VALOR BDI, VALOR COM BDI
                    if any(
                        x in coef_upper
                        for x in ["TOTAL", "VALOR", "VALOR BDI", "VALOR COM BDI"]
                    ):
                        continue
                    # Pular linhas que contêm VALOR, VALOR BDI ou VALOR COM BDI usando valores do JSON
                    if any(x in coef_upper for x in valores_pular):
                        continue

                # ============================================
                # FILTRO POR CÓDIGO DO ITEM (iniciaPor/naoIniciaPor)
                # Aplicar filtro ao código do item (coluna desc)
                # ============================================
                codigo_item = desc.strip()
                codigo_upper = codigo_item.upper()

                # Verificar se iniciaPor está definido e se o código começa com ele
                if inicia_por and not codigo_upper.startswith(inicia_por.upper()):
                    continue

                # Verificar se naoIniciaPor está definido e se o código NÃO começa com ele
                if nao_inicia_por and codigo_upper.startswith(nao_inicia_por.upper()):
                    continue

                # Adicionar fórmula
                if fator_coef:
                    # Adicionar na coluna E (coeficiente)
                    cell_coef = sheet.cell(row=y, column=col_coef_idx)
                    if not isinstance(cell_coef, MergedCell):
                        # NÃO sobrescrever se já tiver uma fórmula
                        if (
                            cell_coef.value
                            and isinstance(cell_coef.value, str)
                            and cell_coef.value.startswith("=")
                        ):
                            pass  # já tem fórmula, não sobrescrever
                        else:
                            # Verificar se a coluna A (código) contém um código válido de item
                            # Códigos válidos começam com dígito OU letra seguida de números (ex: I00378S, S10555, I00081)
                            # Textos descritivos como "Material", "Mão de Obra" não devem receber fórmula
                            cell_item_val = sheet.cell(row=y, column=col_desc_idx).value
                            adicionar_formula = True
                            if cell_item_val:
                                cell_item_str = str(cell_item_val).strip()
                                if cell_item_str:
                                    is_digit_start = cell_item_str[0].isdigit()
                                    is_alpha_with_digits = cell_item_str[
                                        0
                                    ].isalpha() and any(
                                        c.isdigit() for c in cell_item_str
                                    )
                                    if not is_digit_start and not is_alpha_with_digits:
                                        adicionar_formula = False
                            if adicionar_formula:
                                formula = f"={get_column_letter(col_coef_antigo_idx)}{y}*FATOR"
                                cell_coef.value = formula
                                linhas_adicionadas.append(
                                    f"L{y}: {desc[:30]} -> {formula}"
                                )
                else:
                    # Adicionar na coluna F (preço unitário)
                    cell_preco = sheet.cell(row=y, column=col_preco_idx)
                    if not isinstance(cell_preco, MergedCell):
                        if cell_preco.value is None or (
                            not isinstance(cell_preco.value, str)
                            or not cell_preco.value.startswith("=")
                        ):
                            # Verificar se a coluna A (código) contém um código válido de item
                            # Códigos válidos começam com dígito OU letra seguida de números (ex: I00378S, S10555, I00081)
                            # Textos descritivos como "Material", "Mão de Obra" não devem receber fórmula
                            cell_item_val = sheet.cell(row=y, column=col_desc_idx).value
                            adicionar_formula = True
                            if cell_item_val:
                                cell_item_str = str(cell_item_val).strip()
                                if cell_item_str:
                                    is_digit_start = cell_item_str[0].isdigit()
                                    is_alpha_with_digits = cell_item_str[
                                        0
                                    ].isalpha() and any(
                                        c.isdigit() for c in cell_item_str
                                    )
                                    if not is_digit_start and not is_alpha_with_digits:
                                        adicionar_formula = False
                            if adicionar_formula:
                                formula = f"=ROUND({get_column_letter(col_preco_antigo_idx)}{y}*FATOR, 2)"
                                cell_preco.value = formula
                                linhas_adicionadas.append(
                                    f"L{y}: {desc[:30]} -> {formula}"
                                )

        if count_processadas > 0:
            print(f"   Processadas {count_processadas} seções de '{nome_upper}'")

    # Mostrar primeiras linhas adicionadas
    if linhas_adicionadas:
        print(f"   Adicionadas: {len(linhas_adicionadas)}")
        for item in linhas_adicionadas[:10]:
            print(f"      {item}")
        if len(linhas_adicionadas) > 10:
            print(f"      ... e mais {len(linhas_adicionadas) - 10}")

    return len(linhas_adicionadas)
