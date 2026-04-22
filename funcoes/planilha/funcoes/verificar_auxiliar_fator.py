from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.hyperlink import Hyperlink

TEXTOS_SKIP = {
    "MATERIAL",
    "MÃO DE OBRA",
    "SERVIÇO",
    "EQUIPAMENTO",
    "TOTAL",
    "PREÇO",
    "ENCARGOS",
}
TEXTOS_VALOR_SKIP = {"TOTAL", "VALOR", "VALOR BDI", "VALOR COM BDI"}

VALOR_LABEL = "VALOR:"


def _limpar_codigo(codigo):
    """Limpa o código, removendo caracteres especiais."""
    return (
        codigo.replace("\u200b", "").replace("\ufeff", "").strip().split()[0]
        if codigo.replace("\u200b", "").replace("\ufeff", "").strip().split()
        else ""
    )


def _codigo_valido(sheet, row, col_desc_idx):
    """Verifica se a linha contém um código válido."""
    val = sheet.cell(row=row, column=col_desc_idx).value
    if not val:
        return False
    s = str(val).strip()
    if not s:
        return False
    for mr in sheet.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col_desc_idx <= mr.max_col:
            if mr.max_col - mr.min_col > 2:
                return False
    return s[0].isdigit() or (s[0].isalpha() and any(c.isdigit() for c in s))


def _add_hyperlink(sheet, row, col, planilha, linha_ref):
    """Adiciona hyperlink à célula."""
    if linha_ref > 0:
        cell = sheet.cell(row=row, column=col)
        if not isinstance(cell, MergedCell) and not cell.hyperlink:
            cell.hyperlink = Hyperlink(
                ref=cell.coordinate, location=f"'{planilha}'!A{linha_ref}"
            )


def _construir_mapa_mescladas(sheet, col_item_idx, secoes_encontradas=None):
    """Constrói mapa de códigos para hyperlinks usando APENAS células mescladas
    com mais de 3 células. Retorna dict com {codigo: {linha_titulo, linha_valor}}.
    """
    mapa_titulos = {}

    for mr in sheet.merged_cells.ranges:
        # Calcula tamanho da mesclagem
        total_celulas = (mr.max_row - mr.min_row + 1) * (mr.max_col - mr.min_col + 1)

        # Filtrar apenas mesclagens maiores que 3 células
        if total_celulas <= 3:
            continue

        # Verificar se inclui a coluna desejada
        if mr.min_col <= col_item_idx <= mr.max_col:
            val = sheet.cell(row=mr.min_row, column=col_item_idx).value

            if val:
                codigo = _limpar_codigo(str(val))
                if codigo and len(codigo) >= 5:
                    linha_titulo = mr.min_row
                    linha_valor = None

                    # Se temos as seções, buscar o VALOR: correspondente
                    if secoes_encontradas:
                        secao_encontrada = None

                        # Encontrar a seção que contém esta linha
                        for secao in secoes_encontradas:
                            if (
                                secao["linha_inicio"]
                                <= linha_titulo
                                <= secao["linha_fim"]
                            ):
                                secao_encontrada = secao
                                break

                        # Se encontrou seção que contém o código
                        if secao_encontrada:
                            # Buscar VALOR: nesta seção
                            for linha_busca in range(
                                linha_titulo + 1, secao_encontrada["linha_fim"]
                            ):
                                val_e = sheet.cell(row=linha_busca, column=5).value
                                if val_e and VALOR_LABEL.upper() in str(val_e).upper():
                                    linha_valor = linha_busca
                                    break
                            # Se não encontrou na seção atual, buscar nas próximas
                            if not linha_valor:
                                for secao_prox in secoes_encontradas:
                                    if (
                                        secao_prox["linha_inicio"]
                                        > secao_encontrada["linha_inicio"]
                                    ):
                                        for linha_busca in range(
                                            linha_titulo + 1,
                                            secao_prox["linha_fim"],
                                        ):
                                            val_e = sheet.cell(
                                                row=linha_busca, column=5
                                            ).value
                                            if (
                                                val_e
                                                and VALOR_LABEL.upper()
                                                in str(val_e).upper()
                                            ):
                                                linha_valor = linha_busca
                                                break
                                        if linha_valor:
                                            break
                                    if linha_valor:
                                        break
                        else:
                            # Código fora de qualquer seção - buscar na próxima seção
                            for secao_prox in secoes_encontradas:
                                if secao_prox["linha_inicio"] > linha_titulo:
                                    for linha_busca in range(
                                        linha_titulo + 1,
                                        secao_prox["linha_fim"],
                                    ):
                                        val_e = sheet.cell(
                                            row=linha_busca, column=5
                                        ).value
                                        if (
                                            val_e
                                            and VALOR_LABEL.upper()
                                            in str(val_e).upper()
                                        ):
                                            linha_valor = linha_busca
                                            break
                                    if linha_valor:
                                        break

                    mapa_titulos[codigo.upper()] = {
                        "linha_titulo": linha_titulo,
                        "linha_valor": linha_valor,
                    }

    return mapa_titulos


def verificar_auxiliar_fator(workbook, dados, todos_item):
    """Verifica e adiciona fórmulas de fator e hyperlinks em planilhas.

    Fluxo:
    1. Extrai dois mapas do valores_item.json:
       - mapa_nome_inicia: nome, iniciaPor, naoIniciaPor
       - mapa_config: nome, total, adicionarFator, buscarAuxiliar

    2. PROCESSAR COMPOSICOES AUXILIARES primeiro:
       - Durante o FOR, monta mapa_titulos_aux

    3. PROCESSAR COMPOSICOES depois:
       - Usa o mapa construído para criar hyperlinks
    """
    dados_itens = dados[0] if isinstance(dados, list) else dados

    # Extrair itens do JSON
    todos_item_dict = todos_item[0] if isinstance(todos_item, list) else todos_item
    todos_item_data = []
    for key, value in todos_item_dict.items():
        if key.startswith("item") and isinstance(value, dict):
            todos_item_data.append(value)

    # Separar em dois mapas
    mapa_nome_inicia = []
    mapa_config = []

    for item in todos_item_data:
        if not isinstance(item, dict):
            continue
        nome = item.get("nome", "")
        if not nome:
            continue

        mapa_nome_inicia.append(
            {
                "nome": nome,
                "iniciaPor": item.get("iniciaPor", ""),
                "naoIniciaPor": item.get("naoIniciaPor", ""),
            }
        )

        mapa_config.append(
            {
                "nome": nome,
                "total": item.get("total", ""),
                "adicionarFator": item.get("adicionarFator", "Não"),
                "buscarAuxiliar": item.get("buscarAuxiliar", "Não"),
                "fatorCoeficiente": item.get("fatorCoeficiente", "Não") == "Sim",
            }
        )

    # Configurações das planilhas
    planilha_comp = dados_itens.get("planilhaComposicao", "COMPOSICOES")
    planilha_aux = dados_itens.get("planilhaAuxiliar", "COMPOSICOES AUXILIARES")

    col_desc = column_index_from_string(dados_itens.get("composicaoDescricao", "A"))
    col_coef = column_index_from_string(dados_itens.get("composicaoCoeficiente", "E"))
    col_preco = column_index_from_string(
        dados_itens.get("composicaoPrecoUnitario", "F")
    )
    col_coef_antigo = column_index_from_string(
        dados_itens.get("composicaoCoeficienteCopiar", "L")
    )
    col_preco_antigo = column_index_from_string(
        dados_itens.get("composicaoPrecoUnitarioCopiar", "M")
    )

    # Obter worksheets
    sheet_comp = workbook[planilha_comp]
    sheet_aux = workbook[planilha_aux]

    # ==========================================
    # ENCONTRAR TODAS AS SEÇÕES PRIMEIRO
    # ==========================================
    secoes_comp = _encontrar_todas_secoes(sheet_comp, col_desc, mapa_config)
    secoes_aux = _encontrar_todas_secoes(sheet_aux, col_desc, mapa_config)

    # ==========================================
    # CONSTRUIR MAPA DE TÍTULOS DA AUXILIAR
    # Usando células mescladas (apenas títulos principais)
    # Já inclui linha_valor para cada código
    # ==========================================
    mapa_titulos_aux = _construir_mapa_mescladas(sheet_aux, col_desc, secoes_aux)

    # ==========================================
    # PROCESSAR COMPOSICOES AUXILIARES
    # Passa o mapa_titulos_aux para usar linha_valor correta
    # ==========================================
    resultado_aux = _processar_planilha_auxiliar(
        sheet_aux,
        col_desc,
        col_coef,
        col_preco,
        col_coef_antigo,
        col_preco_antigo,
        mapa_nome_inicia,
        mapa_config,
        planilha_aux,
        mapa_titulos_aux,
    )

    # ==========================================
    # PROCESSAR COMPOSICOES DEPOIS
    # Usa o mapa_titulos_aux construído na AUXILIARES
    # ==========================================
    resultado_comp = _processar_planilha(
        sheet_comp,
        col_desc,
        col_coef,
        col_preco,
        col_coef_antigo,
        col_preco_antigo,
        mapa_nome_inicia,
        mapa_config,
        mapa_titulos_aux,
        planilha_aux,
    )

    return {
        "formulas_fator_comp": resultado_comp["formulas_fator"],
        "formulas_fator_aux": resultado_aux["formulas_fator"],
        "formulas_auxiliares_comp": resultado_comp["formulas_auxiliar"],
        "formulas_auxiliares_aux": resultado_aux["formulas_auxiliar"],
        "hyperlinks_criados": resultado_comp["hyperlinks"]
        + resultado_aux["hyperlinks"],
    }


def _processar_planilha(
    sheet,
    col_desc,
    col_coef,
    col_preco,
    col_coef_antigo,
    col_preco_antigo,
    mapa_nome_inicia,
    mapa_config,
    mapa_titulos_aux,
    planilha_aux,
):
    """Processa uma planilha: FOR percorrendo todas as linhas."""
    resultado = {
        "formulas_fator": 0,
        "formulas_auxiliar": 0,
        "hyperlinks": 0,
    }

    max_row = min(sheet.max_row, 20000)
    secao_atual = None

    # Encontrar todas as seções
    secoes_encontradas = _encontrar_todas_secoes(sheet, col_desc, mapa_config)

    # Mapa de busca de códigos na planilha auxiliar
    mapa_busca = {}
    if mapa_titulos_aux:
        mapa_busca = mapa_titulos_aux

    # FOR percorrendo todas as linhas
    for linha in range(1, max_row + 1):
        cell_desc = sheet.cell(row=linha, column=col_desc)

        if isinstance(cell_desc, MergedCell):
            continue

        valor = cell_desc.value
        if not valor:
            continue

        valor_str = str(valor).strip()
        valor_upper = valor_str.upper()

        # Atualizar seção atual
        secao_atual = _verificar_troca_secao(
            linha, valor_upper, secoes_encontradas, secao_atual, mapa_nome_inicia
        )

        if not secao_atual:
            continue

        # Pular linhas com TEXTOS_SKIP
        if any(x in valor_upper for x in TEXTOS_SKIP):
            continue
        if any(x in valor_upper for x in TEXTOS_VALOR_SKIP):
            continue

        # Verificar se está dentro de seção válida
        if linha <= secao_atual["linha_inicio"] or linha >= secao_atual["linha_fim"]:
            continue

        # Verificar se é linha de cabeçalho
        if "COEFICIENTE" in valor_upper or "PREÇO UNITÁRIO" in valor_upper:
            continue

        # Verificar fonte/unidade
        cell_fonte = sheet.cell(row=linha, column=col_desc + 2).value
        cell_unid = sheet.cell(row=linha, column=col_desc + 3).value
        if cell_fonte and "FONTE" in str(cell_fonte).upper():
            continue
        if cell_unid and "UNID" in str(cell_unid).upper():
            continue

        # Verificar se é código válido
        codigo_upper = valor_str.upper()
        if not _codigo_valido(sheet, linha, col_desc):
            continue

        # Verificar filtros
        iniciaPor = secao_atual.get("iniciaPor", "")
        naoIniciaPor = secao_atual.get("naoIniciaPor", "")

        if iniciaPor and not codigo_upper.startswith(iniciaPor.upper()):
            continue
        if naoIniciaPor and codigo_upper.startswith(naoIniciaPor.upper()):
            continue

        # ==========================================
        # ADICIONAR FATOR
        # ==========================================
        if secao_atual.get("adicionarFator"):
            val_coef = sheet.cell(row=linha, column=col_coef).value
            val_preco = sheet.cell(row=linha, column=col_preco).value

            if (
                val_coef and isinstance(val_coef, str) and "*FATOR" in val_coef.upper()
            ) or (
                val_preco
                and isinstance(val_preco, str)
                and "*FATOR" in val_preco.upper()
            ):
                continue

            if secao_atual.get("fatorCoeficiente"):
                cell = sheet.cell(row=linha, column=col_coef)
                if not isinstance(cell, MergedCell):
                    cell.value = f"={get_column_letter(col_coef_antigo)}{linha}*FATOR"
                    resultado["formulas_fator"] += 1
            else:
                cell = sheet.cell(row=linha, column=col_preco)
                if not isinstance(cell, MergedCell):
                    cell.value = (
                        f"=ROUND({get_column_letter(col_preco_antigo)}{linha}*FATOR, 2)"
                    )
                    resultado["formulas_fator"] += 1

        # ==========================================
        # BUSCAR AUXILIAR (criar hyperlink e fórmula)
        # ==========================================
        if secao_atual.get("buscarAuxiliar") and mapa_busca:
            if cell_desc.hyperlink:
                continue

            codigo_limpo = _limpar_codigo(valor_str)
            if codigo_limpo and len(codigo_limpo) >= 5:
                codigo_upper_limpo = codigo_limpo.upper()
                if codigo_upper_limpo in mapa_busca:
                    dados_codigo = mapa_busca[codigo_upper_limpo]
                    linha_titulo = dados_codigo["linha_titulo"]
                    linha_valor = dados_codigo["linha_valor"]

                    # Criar hyperlink
                    _add_hyperlink(
                        sheet,
                        linha,
                        col_desc,
                        planilha_aux,
                        linha_titulo,
                    )
                    resultado["hyperlinks"] += 1

                    # Adicionar fórmula no preço unitário se tiver linha_valor
                    if linha_valor:
                        cell_preco = sheet.cell(row=linha, column=col_preco)
                        if not isinstance(cell_preco, MergedCell):
                            cell_preco.value = f"='{planilha_aux}'!G{linha_valor}"
                            resultado["formulas_auxiliar"] += 1

    return resultado


def _processar_planilha_auxiliar(
    sheet,
    col_desc,
    col_coef,
    col_preco,
    col_coef_antigo,
    col_preco_antigo,
    mapa_nome_inicia,
    mapa_config,
    nome_planilha,
    mapa_titulos_aux=None,
):
    """Processa planilha AUXILIAR: aplica fator e hyperlinks internos."""
    resultado = {
        "formulas_fator": 0,
        "formulas_auxiliar": 0,
        "hyperlinks": 0,
    }

    max_row = min(sheet.max_row, 20000)
    secao_atual = None

    # Encontrar todas as seções
    secoes_encontradas = _encontrar_todas_secoes(sheet, col_desc, mapa_config)

    # FOR percorrendo todas as linhas
    for linha in range(1, max_row + 1):
        cell_desc = sheet.cell(row=linha, column=col_desc)

        if isinstance(cell_desc, MergedCell):
            continue

        valor = cell_desc.value
        if not valor:
            continue

        valor_str = str(valor).strip()
        valor_upper = valor_str.upper()

        # Atualizar seção atual
        secao_atual = _verificar_troca_secao(
            linha, valor_upper, secoes_encontradas, secao_atual, mapa_nome_inicia
        )

        if not secao_atual:
            continue

        # Pular linhas com TEXTOS_SKIP
        if any(x in valor_upper for x in TEXTOS_SKIP):
            continue
        if any(x in valor_upper for x in TEXTOS_VALOR_SKIP):
            continue

        # Verificar se está dentro de seção válida
        if linha <= secao_atual["linha_inicio"] or linha >= secao_atual["linha_fim"]:
            continue

        # Verificar se é linha de cabeçalho
        if "COEFICIENTE" in valor_upper or "PREÇO UNITÁRIO" in valor_upper:
            continue

        # Verificar fonte/unidade
        cell_fonte = sheet.cell(row=linha, column=col_desc + 2).value
        cell_unid = sheet.cell(row=linha, column=col_desc + 3).value
        if cell_fonte and "FONTE" in str(cell_fonte).upper():
            continue
        if cell_unid and "UNID" in str(cell_unid).upper():
            continue

        # Verificar se é código válido
        codigo_upper = valor_str.upper()
        if not _codigo_valido(sheet, linha, col_desc):
            continue

        # Verificar filtros
        iniciaPor = secao_atual.get("iniciaPor", "")
        naoIniciaPor = secao_atual.get("naoIniciaPor", "")

        if iniciaPor and not codigo_upper.startswith(iniciaPor.upper()):
            continue
        if naoIniciaPor and codigo_upper.startswith(naoIniciaPor.upper()):
            continue

        # ==========================================
        # ADICIONAR FATOR
        # ==========================================
        if secao_atual.get("adicionarFator"):
            val_coef = sheet.cell(row=linha, column=col_coef).value
            val_preco = sheet.cell(row=linha, column=col_preco).value

            if (
                val_coef and isinstance(val_coef, str) and "*FATOR" in val_coef.upper()
            ) or (
                val_preco
                and isinstance(val_preco, str)
                and "*FATOR" in val_preco.upper()
            ):
                continue

            if secao_atual.get("fatorCoeficiente"):
                cell = sheet.cell(row=linha, column=col_coef)
                if not isinstance(cell, MergedCell):
                    cell.value = f"={get_column_letter(col_coef_antigo)}{linha}*FATOR"
                    resultado["formulas_fator"] += 1
            else:
                cell = sheet.cell(row=linha, column=col_preco)
                if not isinstance(cell, MergedCell):
                    cell.value = (
                        f"=ROUND({get_column_letter(col_preco_antigo)}{linha}*FATOR, 2)"
                    )
                    resultado["formulas_fator"] += 1

        # ==========================================
        # BUSCAR AUXILIAR (criar hyperlink interno e fórmula)
        # ==========================================
        if secao_atual.get("buscarAuxiliar"):
            if cell_desc.hyperlink:
                continue

            codigo_limpo = _limpar_codigo(valor_str)

            # Criar hyperlink interno (aponta para própria planilha)
            if codigo_limpo and len(codigo_limpo) >= 5:
                codigo_upper = codigo_limpo.upper()

                # Determinar qual linha usar para o hyperlink
                linha_hyperlink = linha  # padrão: própria linha
                linha_valor = None

                # Se temos o mapa de títulos, usar os valores corretos
                if mapa_titulos_aux and codigo_upper in mapa_titulos_aux:
                    dados_codigo = mapa_titulos_aux[codigo_upper]
                    linha_hyperlink = dados_codigo.get("linha_titulo", linha)
                    linha_valor = dados_codigo.get("linha_valor")

                _add_hyperlink(
                    sheet,
                    linha,
                    col_desc,
                    nome_planilha,
                    linha_hyperlink,
                )
                resultado["hyperlinks"] += 1

                # Se temos o linha_valor calculado, adicionar fórmula
                if linha_valor:

                    # Adicionar fórmula apenas se linha_valor está depois do código
                    # e não aponta para a própria seção (evitar referência circular)
                    if linha_valor and linha_valor > linha:
                        cell_preco = sheet.cell(row=linha, column=col_preco)
                        if not isinstance(cell_preco, MergedCell):
                            cell_preco.value = f"=G{linha_valor}"
                            resultado["formulas_auxiliar"] += 1
                else:
                    # Fallback: busca limitada ao fim da seção atual
                    linha_fim_busca = secao_atual["linha_fim"]
                    busca_max = min(linha_fim_busca - 1, linha + 50)

                    for linha_valor in range(linha + 1, busca_max):
                        val_e = sheet.cell(row=linha_valor, column=5).value
                        if val_e and VALOR_LABEL.upper() in str(val_e).upper():
                            cell_preco = sheet.cell(row=linha, column=col_preco)
                            if not isinstance(cell_preco, MergedCell):
                                if linha_valor > linha:
                                    cell_preco.value = f"=G{linha_valor}"
                                    resultado["formulas_auxiliar"] += 1
                            break

    return resultado


def _encontrar_todas_secoes(sheet, col_desc, mapa_config):
    """Encontra todas as seções na planilha (linha_inicio e linha_fim para cada seção)."""
    secoes = []
    max_row = min(sheet.max_row, 20000)
    colunas_busca = list(range(col_desc, col_desc + 10))

    # Primeiro, encontrar TODAS as linhas onde aparece o nome de alguma seção
    linhas_nomes = {}

    for linha in range(1, max_row + 1):
        for col in colunas_busca:
            cell = sheet.cell(row=linha, column=col)
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if val:
                val_limpo = str(val).replace("\u200b", "").replace("\ufeff", "").strip()
                val_upper = val_limpo.upper()

                for config in mapa_config:
                    if val_upper == config["nome"].upper():
                        linhas_nomes.setdefault(config["nome"].upper(), []).append(
                            (linha, config["nome"])
                        )
                        break

    # Também buscar em células mescladas
    for mr in sheet.merged_cells.ranges:
        if mr.min_row > max_row:
            continue
        for col in colunas_busca:
            if mr.min_col <= col <= mr.max_col:
                val = sheet.cell(row=mr.min_row, column=col).value
                if val:
                    val_limpo = (
                        str(val).replace("\u200b", "").replace("\ufeff", "").strip()
                    )
                    val_upper = val_limpo.upper()

                    for config in mapa_config:
                        if val_upper == config["nome"].upper():
                            linhas_nomes.setdefault(config["nome"].upper(), []).append(
                                (mr.min_row, config["nome"])
                            )
                            break

    # Para cada seção encontrada, encontrar o total correspondente
    for nome_upper, ocorrencias in linhas_nomes.items():
        config_encontrada = None
        for config in mapa_config:
            if config["nome"].upper() == nome_upper:
                config_encontrada = config
                break

        if not config_encontrada:
            continue

        total_upper = config_encontrada["total"].upper()

        for linha_inicio, nome_original in ocorrencias:
            # Buscar o total APÓS a linha_inicio
            linha_fim = None
            for linha in range(linha_inicio + 1, max_row + 1):
                for col in colunas_busca:
                    cell = sheet.cell(row=linha, column=col)
                    if isinstance(cell, MergedCell):
                        continue
                    val = cell.value
                    if val:
                        val_limpo = (
                            str(val).replace("\u200b", "").replace("\ufeff", "").strip()
                        )
                        if val_limpo.upper() == total_upper:
                            linha_fim = linha
                            break
                if linha_fim:
                    break

            # Também buscar em células mescladas
            if not linha_fim:
                for mr in sheet.merged_cells.ranges:
                    if mr.min_row <= linha_inicio:
                        continue
                    if mr.min_row > max_row:
                        continue
                    for col in colunas_busca:
                        if mr.min_col <= col <= mr.max_col:
                            val = sheet.cell(row=mr.min_row, column=col).value
                            if val:
                                val_limpo = (
                                    str(val)
                                    .replace("\u200b", "")
                                    .replace("\ufeff", "")
                                    .strip()
                                )
                                if val_limpo.upper() == total_upper:
                                    linha_fim = mr.min_row
                                    break
                    if linha_fim:
                        break

            if linha_fim:
                secoes.append(
                    {
                        "nome": nome_original,
                        "nome_upper": nome_upper,
                        "linha_inicio": linha_inicio,
                        "linha_fim": linha_fim,
                        "adicionarFator": config_encontrada["adicionarFator"] == "Sim",
                        "buscarAuxiliar": config_encontrada["buscarAuxiliar"] == "Sim",
                        "fatorCoeficiente": config_encontrada.get(
                            "fatorCoeficiente", False
                        ),
                    }
                )

    return secoes


def _verificar_troca_secao(
    linha, valor_upper, secoes_encontradas, secao_atual, mapa_nome_inicia
):
    """Verifica se há troca de seção na linha atual."""
    if secao_atual:
        if linha == secao_atual["linha_inicio"]:
            return secao_atual
        if linha >= secao_atual["linha_fim"]:
            secao_atual = None

    # Verificar se entrou em uma nova seção
    for secao in secoes_encontradas:
        if linha == secao["linha_inicio"]:
            iniciaPor = ""
            naoIniciaPor = ""
            for item in mapa_nome_inicia:
                if item["nome"].upper() == secao["nome_upper"]:
                    iniciaPor = item.get("iniciaPor", "")
                    naoIniciaPor = item.get("naoIniciaPor", "")
                    break

            return {
                "nome": secao["nome"],
                "nome_upper": secao["nome_upper"],
                "linha_inicio": secao["linha_inicio"],
                "linha_fim": secao["linha_fim"],
                "adicionarFator": secao["adicionarFator"],
                "buscarAuxiliar": secao["buscarAuxiliar"],
                "fatorCoeficiente": secao["fatorCoeficiente"],
                "iniciaPor": iniciaPor,
                "naoIniciaPor": naoIniciaPor,
            }

    return secao_atual
