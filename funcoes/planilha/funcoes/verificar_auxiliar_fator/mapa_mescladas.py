from openpyxl.cell.cell import MergedCell

from .constantes import VALOR_LABEL, TEXTOS_SKIP
from .limpar import _limpar_codigo


def construir_mapa_mescladas(sheet, col_item_idx, mapa_config=None):
    """Constrói mapa de códigos para hyperlinks.

    Processa:
    1. Identifica títulos (células com >3 colunas mescladas na coluna A)
    2. Extrai código do título (primeira palavra, split por espaço)
    3. Limita busca de VALOR: ao escopo da seção atual (até o próximo título)

    Retorna dict com {codigo: linha_valor}.
    """
    mapa_titulos = {}
    max_row = min(sheet.max_row, 20000)

    # Processar linha por linha
    for linha in range(1, max_row + 1):
        cell = sheet.cell(row=linha, column=col_item_idx)
        val = cell.value

        if not val:
            continue

        # Limpar o valor
        val_str = str(val).replace("\u200b", "").replace("\ufeff", "").strip()

        # Verificar se está em célula mesclada
        merged_range = None
        for mr in sheet.merged_cells.ranges:
            if (
                mr.min_row <= linha <= mr.max_row
                and mr.min_col <= col_item_idx <= mr.max_col
            ):
                merged_range = mr
                break

        num_cols = 0
        if merged_range:
            num_cols = merged_range.max_col - merged_range.min_col + 1

        # Se é título (>3 colunas mescladas), extrair código
        if num_cols > 3:
            # Extrair código do título (split por espaço, primeira parte)
            partes = val_str.split()
            if partes:
                codigo = partes[0]
                codigo_upper = codigo.upper()

                if codigo:
                    # Buscar linha do VALOR: dentro do escopo da seção
                    linha_valor = buscar_linha_valor_escopo(sheet, linha + 1, max_row)

                    # Armazenar no mapa
                    mapa_titulos[codigo_upper] = linha_valor

    return mapa_titulos


def buscar_linha_valor_escopo(sheet, linha_inicio, max_row):
    """Busca linha com 'VALOR:' a partir da linha_inicio até o próximo título."""
    # Primeiro, encontrar o próximo título (>3 colunas mescladas na coluna A)
    proximo_titulo = max_row + 1
    for linha in range(linha_inicio, max_row + 1):
        cell = sheet.cell(row=linha, column=1)
        if cell.value:
            # Verificar se é título
            for mr in sheet.merged_cells.ranges:
                if (
                    mr.min_row <= linha <= mr.max_row
                    and mr.min_col <= 1 <= mr.max_col
                    and mr.max_col - mr.min_col + 1 > 3
                ):
                    proximo_titulo = linha
                    break

        if proximo_titulo != max_row + 1:
            break

    # Buscar VALOR: entre o início e o próximo título
    for linha in range(linha_inicio, min(proximo_titulo, max_row + 1)):
        val_e = sheet.cell(row=linha, column=5).value
        if val_e and VALOR_LABEL.upper() in str(val_e).upper():
            return linha

    return None


def buscar_linha_valor(sheet, linha_inicio, max_row):
    """Busca linha com 'VALOR:' a partir da linha_inicio."""
    for linha in range(linha_inicio + 1, max_row + 1):
        val_e = sheet.cell(row=linha, column=5).value
        if val_e and VALOR_LABEL.upper() in str(val_e).upper():
            return linha
    return None
