from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

from .constantes import TEXTOS_SKIP, TEXTOS_VALOR_SKIP, VALOR_LABEL
from .hyperlink import _add_hyperlink
from .limpar import _codigo_valido, _limpar_codigo


def processar_planilha(
    sheet,
    col_desc,
    col_coef,
    col_preco,
    col_coef_antigo,
    col_preco_antigo,
    mapa_nome_inicia,
    mapa_config,
    planilha_destino,
    mapa_titulos_aux,
):
    """FOR único percorrendo linhas, identificando seções e aplicando regras.

    Identifica:
    - INÍCIO: quando encontra nome da seção (ex: "SERVIÇOS")
    - FIM: quando encontra total da seção (ex: "TOTAL SERVIÇOS:")

    Aplica:
    - adicionarFator: "Sim" → adiciona fórmula com fator
    - buscarAuxiliar: "Sim" → cria hyperlink e fórmula referencing planilha auxiliar

    Suporta múltiplas configs para mesmo nome de seção (ex: dois "Serviço" com filtros diferentes).
    """
    resultado = {
        "formulas_fator": 0,
        "formulas_auxiliar": 0,
        "hyperlinks": 0,
    }

    max_row = min(sheet.max_row, 20000)

    # Estado da seção atual
    secao_atual = None
    linha_inicio_secao = None
    linha_fim_secao = None
    configs_secao = []  # Lista de configs para a seção atual (suporta múltiplas)

    # Mapa de códigos por seção (evita sobrescrever códigos de outras seções)
    mapa_secao_atual = {}

    # FOR único percorrendo todas as linhas
    for linha in range(1, max_row + 1):
        cell_desc = sheet.cell(row=linha, column=col_desc)

        if isinstance(cell_desc, MergedCell):
            continue

        valor = cell_desc.value
        if not valor:
            continue

        valor_str = str(valor).replace("\u200b", "").replace("\ufeff", "").strip()
        valor_upper = valor_str.upper()

        # ==========================================
        # IDENTIFICAR INÍCIO DE SEÇÃO (pelo nome)
        # Coleta TODAS as configs com mesmo nome (suporta item11 e item18 ambos "Serviço")
        # ==========================================
        configs_encontradas = []
        for config in mapa_config:
            if valor_upper == config["nome"].upper():
                configs_encontradas.append(config)

        if configs_encontradas:
            secao_atual = configs_encontradas[0]["nome"]
            linha_inicio_secao = linha
            configs_secao = configs_encontradas  # Armazena TODAS as configs
            # Reset mapa da seção para evitar contaminação de seções anteriores
            mapa_secao_atual = {}
            continue

        # ==========================================
        # IDENTIFICAR FIM DE SEÇÃO (pelo total)
        # Usa a primeira config para determinar o total
        # ==========================================
        if secao_atual and configs_secao:
            config_primeira = configs_secao[0]
            total_upper = config_primeira.get("total", "").upper()
            if valor_upper == total_upper:
                linha_fim_secao = linha
                # Reset para próxima seção
                secao_atual = None
                linha_inicio_secao = None
                linha_fim_secao = None  # IMPORTANTE: reset para permitir próxima seção
                configs_secao = []
                mapa_secao_atual = {}
                continue

        # ==========================================
        # PROCESSAR LINHAS DENTRO DA SEÇÃO
        # ==========================================
        if secao_atual and configs_secao and linha > linha_inicio_secao:
            # Verificar se não mudou de seção
            if linha_fim_secao and linha >= linha_fim_secao:
                secao_atual = None
                linha_inicio_secao = None
                configs_secao = []
                mapa_secao_atual = {}
                continue

            # Pular linhas com TEXTOS_SKIP
            if any(x in valor_upper for x in TEXTOS_SKIP):
                continue
            if any(x in valor_upper for x in TEXTOS_VALOR_SKIP):
                continue

            # Pular linhas de cabeçalho
            if "COEFICIENTE" in valor_upper or "PREÇO UNITÁRIO" in valor_upper:
                continue

            # Verificar fonte/unidade
            cell_fonte = sheet.cell(row=linha, column=col_desc + 2).value
            cell_unid = sheet.cell(row=linha, column=col_desc + 3).value
            if cell_fonte and "FONTE" in str(cell_fonte).upper():
                continue
            if cell_unid and "UNID" in str(cell_unid).upper():
                continue

            # Verificar se é código válido
            if not _codigo_valido(sheet, linha, col_desc):
                continue

            # ==========================================
            # ENCONTRAR CONFIG ESPECÍFICA PARA ESTE CÓDIGO
            # Baseado nos filtros iniciaPor/naoIniciaPor
            # Prioridade: itens com filtros (iniciaPor/naoIniciaPor) são verificados primeiro
            # ==========================================
            config_especifica = None
            melhor_prioridade = -1  # -1 = sem filtro, 0 = com filtro, 1 = não aplicável

            for item in mapa_nome_inicia:
                if item["nome"].upper() != secao_atual.upper():
                    continue

                ip = item.get("iniciaPor", "")
                nip = item.get("naoIniciaPor", "")

                # Verifica naoIniciaPor primeiro (exclusão)
                if nip and valor_str.upper().startswith(nip.upper()):
                    continue  # Match na exclusion, pula

                # Verifica iniciaPor
                if ip:
                    if not valor_str.upper().startswith(ip.upper()):
                        continue  # Não match, pula
                    # Código match com filtro - alta prioridade
                    prioridade = 0
                else:
                    # Sem filtro - baixa prioridade (usar só se não houver match melhor)
                    prioridade = -1

                # Encontrar config correspondente para este item
                for cfg in configs_secao:
                    if cfg["nome"].upper() == item["nome"].upper():
                        # Seleciona a config com melhor prioridade
                        # (prioridade menor = mais específico)
                        if prioridade < melhor_prioridade or melhor_prioridade == -1:
                            config_especifica = cfg
                            melhor_prioridade = prioridade
                        break

            # Se não encontrou config específica, usa a primeira (comportamento original)
            if not config_especifica:
                config_especifica = configs_secao[0]

            # ==========================================
            # ADICIONAR FATOR (usando config específica)
            # ==========================================
            if config_especifica:
                resultado["formulas_fator"] += adicionar_fator(
                    sheet,
                    linha,
                    col_coef,
                    col_preco,
                    col_coef_antigo,
                    col_preco_antigo,
                    config_especifica,
                    secao_atual,
                )

            # ==========================================
            # BUSCAR AUXILIAR (usando config específica)
            # ==========================================
            if config_especifica and config_especifica.get("buscarAuxiliar") == "Sim":
                resultado_busca = buscar_auxiliar(
                    sheet,
                    linha,
                    col_desc,
                    col_preco,
                    valor_str,
                    planilha_destino,
                    mapa_titulos_aux,
                )
                resultado["hyperlinks"] += resultado_busca["hyperlinks"]
                resultado["formulas_auxiliar"] += resultado_busca["formulas_auxiliar"]

    return resultado


def adicionar_fator(
    sheet,
    linha,
    col_coef,
    col_preco,
    col_coef_antigo,
    col_preco_antigo,
    config_secao,
    secao_atual,
):
    """Adiciona fórmula de fator na linha."""
    count = 0
    if config_secao.get("adicionarFator") == "Sim":
        if config_secao.get("fatorCoeficiente"):
            cell = sheet.cell(row=linha, column=col_coef)
            if not isinstance(cell, MergedCell):
                # NÃO sobrescrever se já tiver hyperlink (evita referência circular)
                if not sheet.cell(row=linha, column=1).hyperlink:
                    cell.value = f"={get_column_letter(col_coef_antigo)}{linha}*FATOR"
                    count = 1
        else:
            cell = sheet.cell(row=linha, column=col_preco)
            if not isinstance(cell, MergedCell):
                # NÃO sobrescrever se já tiver hyperlink (evita referência circular)
                if not sheet.cell(row=linha, column=1).hyperlink:
                    cell.value = (
                        f"=ROUND({get_column_letter(col_preco_antigo)}{linha}*FATOR, 2)"
                    )
                    count = 1

    return count


def buscar_auxiliar(
    sheet,
    linha,
    col_desc,
    col_preco,
    valor_str,
    planilha_destino,
    mapa_titulos_aux,
):
    """Busca código no mapa e cria hyperlink/fórmula."""
    resultado = {"hyperlinks": 0, "formulas_auxiliar": 0}

    cell_desc = sheet.cell(row=linha, column=col_desc)
    if cell_desc.hyperlink:
        return resultado

    codigo_limpo = _limpar_codigo(valor_str)
    if codigo_limpo:
        codigo_upper = codigo_limpo.upper()

        # Buscar no mapa pré-construído (mapa_titulos_aux)
        if mapa_titulos_aux and codigo_upper in mapa_titulos_aux:
            linha_valor = mapa_titulos_aux[codigo_upper]

            if linha_valor:
                # Criar hyperlink
                _add_hyperlink(sheet, linha, col_desc, planilha_destino, linha_valor)
                resultado["hyperlinks"] = 1

                # Adicionar fórmula no preço unitário (referenciando a aba auxiliar)
                cell_preco = sheet.cell(row=linha, column=col_preco)
                if not isinstance(cell_preco, MergedCell):
                    cell_preco.value = f"='{planilha_destino}'!G{linha_valor}"
                    resultado["formulas_auxiliar"] = 1

    return resultado


def buscar_linha_valor(sheet, linha_inicio, max_row):
    """Busca linha com 'VALOR:' a partir da linha_inicio."""
    for linha in range(linha_inicio + 1, max_row + 1):
        val_e = sheet.cell(row=linha, column=5).value
        if val_e and VALOR_LABEL.upper() in str(val_e).upper():
            return linha
    return None
