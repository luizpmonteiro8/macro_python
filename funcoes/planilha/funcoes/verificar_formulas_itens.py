from openpyxl.utils import column_index_from_string
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.hyperlink import Hyperlink


def verificar_e_adicionar_formulas(workbook, dados):
    """
    Verifica e adiciona fórmulas para itens com 'buscarAuxiliar': 'Sim'.
    Para células mescladas com código, busca a próxima linha com "VALOR:" para referência.
    Aplica filtros iniciaPor/naoIniciaPor ao CÓDIGO do item.
    Também cria hyperlinks na coluna de descrição (B) apontando para o título da seção.
    """
    print(">>> Verificando fórmulas dos itens auxiliares...")

    sheet = workbook[dados.get("planilhaAuxiliar", "COMPOSICOES AUXILIARES")]
    col_item = dados.get("auxiliarDescricao", "A")  # Coluna do código
    col_preco = dados.get("auxiliarPrecoUnitario", "F")  # Coluna F
    val_str = dados.get("valor", "VALOR:")  # String "VALOR:"
    planilha_aux = dados.get("planilhaAuxiliar", "COMPOSICOES AUXILIARES")

    col_item_idx = column_index_from_string(col_item)
    col_preco_idx = column_index_from_string(col_preco)
    col_desc_idx = 2  # Coluna B - descrição

    max_row = sheet.max_row
    linhas_modificadas = []

    # ============================================
    # Construir lista de itens com buscarAuxiliar: "Sim" e seus filtros
    # ============================================
    dados_itens = dados
    if isinstance(dados, list) and len(dados) > 0:
        dados_itens = dados[0]

    itens_auxiliares = []
    for key, item in dados_itens.items():
        if key.startswith("item") and isinstance(item, dict):
            if item.get("buscarAuxiliar") == "Sim":
                itens_auxiliares.append(
                    {
                        "nome": item.get("nome", ""),
                        "total": item.get("total", ""),
                        "iniciaPor": item.get("iniciaPor", ""),
                        "naoIniciaPor": item.get("naoIniciaPor", ""),
                    }
                )

    print(f">> Itens com buscarAuxiliar: Sim - {len(itens_auxiliares)}")

    # ============================================
    # Primeiro: construir mapa de códigos -> linha do título (célula mesclada)
    # ✅ CORRIGIDO: Usar APENAS codigo_completo para evitar conflitos
    # "88316" e "S88316S" são itens diferentes e devem ser encontrados corretamente
    # ============================================
    mapa_codigos_titulo = {}

    merged_ranges = list(sheet.merged_cells.ranges)
    for merged_range in merged_ranges:
        if merged_range.min_col <= col_item_idx <= merged_range.max_col:
            cell_val = sheet.cell(row=merged_range.min_row, column=col_item_idx).value
            if cell_val:
                codigo = str(cell_val).strip()
                codigo_limpo = (
                    codigo.replace("\u200b", "").replace("\ufeff", "").strip()
                )
                # Pegar apenas o primeiro elemento (código numérico)
                codigo_limpo = codigo_limpo.split()[0] if codigo_limpo.split() else ""
                codigo_completo = codigo_limpo.upper()

                if len(codigo_completo) >= 5:
                    mapa_codigos_titulo[codigo_completo] = merged_range.min_row

    # ============================================
    # Segundo: mapa de códigos -> linha de "VALOR:"
    # ✅ CORRIGIDO: Usar APENAS codigo_completo para evitar conflitos
    # ============================================
    mapa_codigos_valor = {}

    for merged_range in merged_ranges:
        if merged_range.min_col <= col_item_idx <= merged_range.max_col:
            cell_val = sheet.cell(row=merged_range.min_row, column=col_item_idx).value
            if cell_val:
                codigo = str(cell_val).strip()
                codigo_limpo = (
                    codigo.replace("\u200b", "").replace("\ufeff", "").strip()
                )
                # Pegar apenas o primeiro elemento (código numérico)
                codigo_limpo = codigo_limpo.split()[0] if codigo_limpo.split() else ""
                codigo_completo = codigo_limpo.upper()

                if len(codigo_completo) >= 5:
                    linha_valor = -1
                    for j in range(merged_range.min_row + 1, max_row + 1):
                        cell_e = sheet.cell(row=j, column=5).value
                        if cell_e and val_str.upper() in str(cell_e).upper():
                            linha_valor = j
                            break

                    if linha_valor > 0:
                        mapa_codigos_valor[codigo_completo] = linha_valor

    print(f">> Códigos com referência a 'VALOR:': {len(mapa_codigos_valor)}")

    # ============================================
    # Processar cada linha
    # ============================================
    hyperlinks_criados = 0

    for i in range(1, max_row + 1):
        cell_item = sheet.cell(row=i, column=col_item_idx).value
        if not cell_item or isinstance(cell_item, MergedCell):
            continue

        codigo = str(cell_item).strip()
        if any(
            x in codigo.upper()
            for x in [
                "MATERIAL",
                "MÃO DE OBRA",
                "SERVIÇO",
                "EQUIPAMENTO",
                "TOTAL",
                "PREÇO",
                "ENCARGOS",
            ]
        ):
            continue

        cell_f = sheet.cell(row=i, column=col_preco_idx).value
        if isinstance(cell_f, MergedCell):
            continue

        tem_formula = cell_f and isinstance(cell_f, str) and cell_f.startswith("=")

        codigo_limpo = codigo.replace("\u200b", "").replace("\ufeff", "").strip()
        # Pegar apenas o primeiro elemento (código numérico)
        codigo_limpo = codigo_limpo.split()[0] if codigo_limpo.split() else ""
        codigo_upper = codigo_limpo.upper()
        codigo_completo = codigo_limpo.upper()

        # ============================================
        # Criar hyperlink para itens com fórmula existente
        # ✅ CORRIGIDO: Verificar se o código é válido antes de procurar na mapa
        # ============================================
        if tem_formula:
            # Verificar se o código é válido (não é texto como "Material", "Mão de Obra", etc)
            codigo_valido = codigo_completo and len(codigo_completo) >= 5

            linha_titulo = -1
            if codigo_valido and codigo_completo in mapa_codigos_titulo:
                linha_titulo = mapa_codigos_titulo[codigo_completo]

            if linha_titulo > 0:
                cell_desc = sheet.cell(row=i, column=col_desc_idx)
                if not isinstance(cell_desc, MergedCell) and not cell_desc.hyperlink:
                    cell_desc.hyperlink = Hyperlink(
                        ref=cell_desc.coordinate,
                        location=f"'{planilha_aux}'!A{linha_titulo}",
                    )
                    hyperlinks_criados += 1
            continue

        # ============================================
        # Filtro por código do item
        # ============================================
        codigo_aprovado = False

        # Primeiro: verificar se algum item tem filtros específicos definidos
        algum_filtro_especifico = any(
            item.get("iniciaPor") or item.get("naoIniciaPor")
            for item in itens_auxiliares
        )

        if algum_filtro_especifico:
            # Se algum item tem filtro específico, verificar se este código corresponde
            for item_info in itens_auxiliares:
                inicia_por = item_info.get("iniciaPor", "")
                nao_inicia_por = item_info.get("naoIniciaPor", "")

                # Ignorar itens com filtros vazios - não devem aprovar por padrão
                if not inicia_por and not nao_inicia_por:
                    continue

                inicia_ok = True
                if inicia_por and not codigo_upper.startswith(inicia_por.upper()):
                    inicia_ok = False

                nao_ok = True
                if nao_inicia_por and codigo_upper.startswith(nao_inicia_por.upper()):
                    nao_ok = False

                if inicia_ok and nao_ok:
                    codigo_aprovado = True
                    break
        else:
            # Se nenhum item tem filtro específico, aprovar (comportamento original)
            codigo_aprovado = True

        if not codigo_aprovado:
            continue

        # ============================================
        # Adicionar fórmula e hyperlink
        # ✅ CORRIGIDO: Verificar se o código é válido (>= 5 caracteres) antes de adicionar
        # ============================================
        chave_encontrada = None
        if (
            codigo_completo
            and len(codigo_completo) >= 5
            and codigo_completo in mapa_codigos_valor
        ):
            chave_encontrada = codigo_completo

        if chave_encontrada:
            ref_linha = mapa_codigos_valor[chave_encontrada]
            if ref_linha != i:
                cell_destino = sheet.cell(row=i, column=col_preco_idx)
                if not isinstance(cell_destino, MergedCell):
                    cell_destino.value = f"=G{ref_linha}"
                    linhas_modificadas.append((i, codigo_limpo[:50], ref_linha))

                    # Criar hyperlink
                    linha_titulo = -1
                    if chave_encontrada in mapa_codigos_titulo:
                        linha_titulo = mapa_codigos_titulo[chave_encontrada]

                    if linha_titulo > 0:
                        cell_desc = sheet.cell(row=i, column=col_desc_idx)
                        if not isinstance(cell_desc, MergedCell):
                            cell_desc.hyperlink = Hyperlink(
                                ref=cell_desc.coordinate,
                                location=f"'{planilha_aux}'!A{linha_titulo}",
                            )
                            hyperlinks_criados += 1

    print(f">> Fórmulas adicionadas: {len(linhas_modificadas)}")
    print(f">> Hyperlinks criados: {hyperlinks_criados}")

    return len(linhas_modificadas)
