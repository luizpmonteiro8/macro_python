"""
Componente para validar o nome de uma planilha.
Inclui tratamento correto do botão Cancelar.
"""

from openpyxl.utils import column_index_from_string

from funcoes.validar.janela_corrigir import janela_corrigir_valor
from funcoes.validar.salvar_json import salvar_json_corrigido


def validar_nome_planilha(
    workbook,
    nome_planilha,
    nome_exibicao,
    erros,
    dados,
    indice_config,
    abas_disponiveis,
):
    """
    Valida que o nome da planilha não está vazio e existe no arquivo.
    TRATAMENTO CORRETO: Se usuário cancelar, retorna False imediatamente.

    Args:
        workbook: Objeto workbook do openpyxl
        nome_planilha: Nome da planilha a validar
        nome_exibicao: Nome para exibição nas mensagens de erro
        erros: Lista de erros
        dados: Dados do JSON (para correção)
        indice_config: Índice da configuração
        abas_disponiveis: Lista de abas disponíveis no Excel

    Returns:
        tuple: (válido: bool, sheet: Worksheet ou None, corrigido: bool)
    """
    if nome_planilha is None or nome_planilha.strip() == "":
        # LOOP: continuar pedindo até acertar ou cancelar
        while True:
            instrucao = (
                "1. Abra o arquivo Excel\n"
                "2. Observe o nome da aba na parte inferior da tela\n"
                "3. Clique na aba que deseja usar\n"
                "4. Digite o nome exatamente como aparece"
            )
            confirmado, novo_valor = janela_corrigir_valor(
                titulo="Nome da aba não definido",
                mensagem=f"O nome da aba '{nome_exibicao}' não foi definido nas configurações.",
                instrucao=instrucao,
                valor_atual="",
                valor_default="PLANILHA ORCAMENTARIA",
            )

            # Se cancelou, para
            if not confirmado:
                return False, None, False

            # Se digitou algo
            if novo_valor:
                # Verifica se existe
                if novo_valor in workbook.sheetnames:
                    dados[indice_config][
                        (
                            "planilhaOrcamentaria"
                            if "Orçament" in nome_exibicao
                            else (
                                "planilhaFator"
                                if "Resumo" in nome_exibicao
                                else (
                                    "planilhaComposicao"
                                    if "Composi" in nome_exibicao
                                    else "planilhaAuxiliar"
                                )
                            )
                        )
                    ] = novo_valor
                    salvar_json_corrigido(dados, indice_config)
                    return True, None, True
                else:
                    # Não existe - mostra erro e CONTINUA o loop
                    from tkinter import messagebox

                    messagebox.showerror(
                        "Erro",
                        f"A aba '{novo_valor}' não existe no arquivo! Tente novamente.",
                    )
                    continue  # Continua o loop para tentar novamente

    if nome_planilha not in workbook.sheetnames:
        lista_abas = (
            ", ".join(workbook.sheetnames) if workbook.sheetnames else "nenhuma"
        )

        # LOOP: continuar pedindo até acertar ou cancelar
        while True:
            instrucao = (
                f"1. Abra o arquivo Excel\n"
                f"2. Observe as abas na parte inferior da tela\n"
                f"3. As abas disponíveis são: {lista_abas}\n"
                f"4. Digite o nome correto da aba"
            )
            confirmado, novo_valor = janela_corrigir_valor(
                titulo="Aba não encontrada",
                mensagem=f"A aba '{nome_exibicao}' não foi encontrada no arquivo Excel.",
                instrucao=instrucao,
                valor_atual=nome_planilha,
                valor_default=(
                    lista_abas.split(",")[0].strip() if "," in lista_abas else ""
                ),
            )

            # Se cancelou, para
            if not confirmado:
                return False, None, False

            # Se digitou algo
            if novo_valor:
                # Verifica se existe
                if novo_valor in workbook.sheetnames:
                    # Atualiza e salva
                    if "Orçament" in nome_exibicao:
                        dados[indice_config]["planilhaOrcamentaria"] = novo_valor
                    elif "Resumo" in nome_exibicao:
                        dados[indice_config]["planilhaFator"] = novo_valor
                    elif "Composi" in nome_exibicao:
                        dados[indice_config]["planilhaComposicao"] = novo_valor
                    elif "Auxili" in nome_exibicao:
                        dados[indice_config]["planilhaAuxiliar"] = novo_valor
                    salvar_json_corrigido(dados, indice_config)
                    return True, None, True
                else:
                    # Não existe - mostra erro e CONTINUA o loop
                    from tkinter import messagebox

                    messagebox.showerror(
                        "Erro",
                        f"A aba '{novo_valor}' não existe no arquivo! Tente novamente.",
                    )
                    continue  # Continua o loop para tentar novamente

    return True, workbook[nome_planilha], False
