"""
Componente para validar a planilha orçamentária com correção interativa.
"""

import json

from openpyxl.utils import column_index_from_string

from funcoes.common.buscar_palavras import buscar_palavra
from funcoes.get.get_linhas_json import get_valor_inicial, get_valor_final
from funcoes.validar.janela_corrigir import janela_corrigir_valor
from funcoes.validar.salvar_json import salvar_json_corrigido
from funcoes.validar.validar_coluna import validar_coluna_existe
from funcoes.validar.validar_nome_planilha import validar_nome_planilha
from funcoes.validar.validar_valor_existe_na_coluna import validar_valor_existe_na_coluna


CAMINHO_JSON = "config/valores_colunas.json"


def validar_planilha_orcamentaria(workbook, dados, erros, indice_config=0):
    """
    Valida a planilha orçamentária com correção interativa.

    Args:
        workbook: Objeto workbook do openpyxl
        dados: Dicionário com configurações do arquivo JSON
        erros: Lista de erros
        indice_config: Índice da configuração no JSON

    Returns:
        tuple: (válido: bool, sheet: Worksheet ou None, linha_cabecalhos: int ou None)
    """
    nome_planilha = dados[indice_config].get(
        "planilhaOrcamentaria", "PLANILHA ORCAMENTARIA"
    )

    valido, sheet, corrigido = validar_nome_planilha(
        workbook,
        nome_planilha,
        "Orçamentária",
        erros,
        dados,
        indice_config,
        workbook.sheetnames if hasattr(workbook, "sheetnames") else [],
    )
    if not valido:
        return False, None, None

    if sheet and sheet.max_row < 2:
        erros.append(
            f"ERRO: A aba '{nome_planilha}' está vazia ou não tem dados suficientes."
        )
        return False, sheet, None

    dados = json.load(open(CAMINHO_JSON, "r", encoding="utf-8"))

    coluna_inicial = dados[indice_config].get("colunaInicial", "A")
    valor_inicial = dados[indice_config].get("valorInicial", "ITEM")
    valor_final = dados[indice_config].get("valorFinal", "VALOR BDI TOTAL")

    if sheet:
        if not validar_coluna_existe(
            sheet,
            coluna_inicial,
            "Inicial",
            erros,
            dados,
            indice_config,
            nome_planilha,
            "colunaInicial",
        ):
            return False, sheet, None

        coluna_final = dados[indice_config].get("colunaFinal", "F")
        if not validar_coluna_existe(
            sheet,
            coluna_final,
            "Coluna Final",
            erros,
            dados,
            indice_config,
            nome_planilha,
            "colunaFinal",
        ):
            return False, sheet, None

        linha_cabecalhos = buscar_palavra(sheet, coluna_inicial, valor_inicial)

        if linha_cabecalhos == -1:
            instrucao_coluna = (
                f"1. Abra o arquivo Excel\n"
                f"2. Vá até a aba '{nome_planilha}'\n"
                f"3. Observe as letras no topo das colunas (A, B, C, D...)\n"
                f"4. Digite a letra da coluna que contém '{valor_inicial}' ou 'ITEM'"
            )
            confirmado_coluna, nova_coluna = janela_corrigir_valor(
                titulo="Coluna do valor inicial",
                mensagem=f"O valor '{valor_inicial}' não foi encontrado na coluna '{coluna_inicial}'.\n"
                f"Verifique se a COLUNA está correta!",
                instrucao=instrucao_coluna,
                valor_atual=coluna_inicial,
                valor_default="A",
            )

            if confirmado_coluna and nova_coluna:
                try:
                    column_index_from_string(nova_coluna.upper())
                    dados[indice_config]["colunaInicial"] = nova_coluna.upper()
                    salvar_json_corrigido(dados, indice_config)
                    coluna_inicial = nova_coluna.upper()
                except Exception:
                    from tkinter import messagebox
                    messagebox.showerror("Erro", f"Coluna '{nova_coluna}' inválida!")
                    return False, sheet, None

            instrucao_valor = (
                f"1. Abra o arquivo Excel\n"
                f"2. Vá até a aba '{nome_planilha}'\n"
                f"3. Vá até a coluna '{coluna_inicial}' e procure pelo texto 'ITEM' ou cabeçalho inicial\n"
                f"4. Digite o texto EXATAMENTE como aparece na célula"
            )
            confirmado_valor, novo_valor = janela_corrigir_valor(
                titulo="Texto inicial não encontrado",
                mensagem=f"O texto '{valor_inicial}' não foi encontrado na coluna '{coluna_inicial}'.\n"
                f"Digite o texto correto que aparece no Excel.",
                instrucao=instrucao_valor,
                valor_atual=valor_inicial,
                valor_default="ITEM",
            )

            if confirmado_valor and novo_valor:
                dados[indice_config]["valorInicial"] = novo_valor
                salvar_json_corrigido(dados, indice_config)

                linha_cabecalhos = buscar_palavra(sheet, coluna_inicial, novo_valor)
                if linha_cabecalhos == -1:
                    erros.append(
                        f"ERRO: O texto '{novo_valor}' ainda não foi encontrado na coluna '{coluna_inicial}'."
                    )
                    return False, sheet, None
            else:
                return False, sheet, None

        valores_linha = []
        for cell in sheet[linha_cabecalhos + 1]:
            if cell.value is not None:
                valores_linha.append(str(cell.value).strip().upper())
            else:
                valores_linha.append("")

        cabecalhos_esperados = ["ITEM", "CÓDIGO", "DESCRIÇÃO", "UND", "QUANTIDADE"]

        cabecalhos_faltantes = []
        for cabecalho in cabecalhos_esperados:
            encontrado = False
            for valor in valores_linha:
                if cabecalho.upper() in valor or valor in cabecalho.upper():
                    encontrado = True
                    break
            if not encontrado:
                cabecalhos_faltantes.append(cabecalho)

        if len(cabecalhos_faltantes) > 0:
            erros.append(
                f"ERRO: Algumas colunas obrigatórias não foram encontradas na linha {linha_cabecalhos + 1} da aba '{nome_planilha}'.\n"
                f"Colunas esperadas: {', '.join(cabecalhos_esperados)}\n"
                f"Colunas encontradas: {', '.join(valores_linha)}\n"
                f"Faltando: {', '.join(cabecalhos_faltantes)}"
            )
            return False, sheet, linha_cabecalhos

        if valor_final:
            if not validar_valor_existe_na_coluna(
                sheet,
                coluna_final,
                valor_final,
                "valor_total",
                nome_planilha,
                erros,
                dados,
                indice_config,
                "valorFinal",
            ):
                return False, sheet, linha_cabecalhos

        return True, sheet, linha_cabecalhos

    return True, sheet, None
