"""
Script para processar o arquivo Excel original usando as funções da macro.
Este script simula o que a macro faz, mas de forma direta.
"""

import sys
import os
import time
import shutil

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook
import openpyxl

# Importar funções da macro
from config.open_config import open_valores_colunas, open_valores_item
from funcoes.common.buscar_palavras import buscar_palavra
from funcoes.planilha.funcoes.verificar_auxiliar_fator import verificar_auxiliar_fator
from funcoes.planilha.funcoes.criar_hiperlinks_auxiliares import (
    criar_hiperlinks_auxiliares,
)
from funcoes.common.custo_unitario import custo_unitario_execucao
from funcoes.planilha.funcoes.resume import resumo_totais
from funcoes.planilha.funcoes.adicionar_bdi import adicionar_bdi
from funcoes.planilha.funcoes.adicionar_fator import adicionar_fator
from funcoes.planilha.funcoes.adicionar_fator_aux import adicionar_fator_aux
from funcoes.planilha.funcoes.adicionar_fator_comp import adicionar_fator_comp
from funcoes.planilha.funcoes.criar_hiperlinks_composicao import (
    criar_hiperlinks_composicao,
)
from funcoes.planilha.funcoes.formula_planilha import (
    copiar_coluna_planilha,
    formula_planilha,
)
from funcoes.get.get_linhas_json import (
    get_coluna_final,
    get_coluna_inicial,
    get_planilha_orcamentaria,
    get_valor_final,
    get_valor_inicial,
)


def processar_arquivo(caminho_entrada, caminho_saida, dados, todos_item):
    """Processa o arquivo Excel como a macro faria."""
    print("=" * 60)
    print("PROCESSAMENTO DO ARQUIVO EXCEL")
    print("=" * 60)

    start_time = time.time()

    # Carregar workbook
    print("\n>>> Carregando arquivo Excel...")
    workbook = load_workbook(caminho_entrada)

    # Obter nome da planilha orçamentária
    print(">>> Obtendo nome da planilha orçamentária...")
    sheet_name = get_planilha_orcamentaria(dados)
    print(f"    Planilha: {sheet_name}")
    sheet_planilha = workbook[sheet_name]

    # Obter colunas e valores
    print(">>> Obtendo colunas e valores do JSON...")
    coluna_inicial = get_coluna_inicial(dados)
    valor_inicial = get_valor_inicial(dados)
    coluna_final = get_coluna_final(dados)
    valor_final = get_valor_final(dados)

    # Buscar linhas
    print(">>> Buscando linha inicial...")
    linhaIni = buscar_palavra(sheet_planilha, coluna_inicial, valor_inicial) + 1
    print(f"    linhaIni = {linhaIni}")

    print(">>> Buscando linha final...")
    linhafinal = buscar_palavra(sheet_planilha, coluna_final, valor_final)
    print(f"    linhaFinal = {linhafinal}")
    if linhafinal == -1:
        linhafinal = sheet_planilha.max_row

    # Processamentos
    print("\n>>> Copiando colunas da planilha...")
    copiar_coluna_planilha(sheet_planilha, dados)

    print(">>> Adicionando Fator...")
    adicionar_fator(workbook, dados)

    print(">>> Adicionando BDI...")
    adicionar_bdi(workbook, dados)

    print(">>> Inserindo fórmulas na planilha...")
    formula_planilha(workbook, linhaIni, linhafinal, dados)

    print(">>> Adicionando Fator Auxiliar...")
    adicionar_fator_aux(workbook, dados)

    print(">>> Adicionando Fator de Composição...")
    adicionar_fator_comp(workbook, dados)

    print(">>> Criando hyperlinks com COMPOSICOES...")
    criar_hiperlinks_composicao(workbook, dados, linhaIni, linhafinal)

    print(">>> Criando hyperlinks para itens auxiliares...")
    criar_hiperlinks_auxiliares(workbook, dados, todos_item)

    print(">>> Calculando custo unitário de execução...")
    custo_unitario_execucao(workbook, dados)

    print(">>> Gerando resumo de totais...")
    resumo_totais(workbook, dados)

    # Verificações finais
    print("\n>>> Verificando fórmulas e fatores dos itens...")
    resultados = verificar_auxiliar_fator(workbook, dados, todos_item)
    print(f"    Fórmulas fator (COMPOSICOES): {resultados['formulas_fator_comp']}")
    print(f"    Fórmulas fator (AUXILIARES): {resultados['formulas_fator_aux']}")
    print(f"    Hyperlinks criados: {resultados['hyperlinks_criados']}")

    # Salvar
    print("\n>>> Salvando arquivo processado...")
    workbook.save(caminho_saida)

    tempo_total = time.time() - start_time
    print(f"\n✓ ARQUIVO GERADO COM SUCESSO!")
    print(f"  Arquivo de entrada: {caminho_entrada}")
    print(f"  Arquivo de saída: {caminho_saida}")
    print(f"  Tempo total: {tempo_total:.2f} segundos")

    return resultados


def main():
    # Caminhos
    caminho_entrada = r"testar\orcamento-completo.xlsx"
    caminho_saida = r"testar\excel-final\orcamento-processado-final.xlsx"

    # Carregar configuração
    print("\n>>> Carregando configuração...")
    todos_dados = open_valores_colunas()
    todos_item = open_valores_item()

    dados = todos_dados[0]  # Usar primeira configuração

    # Verificar se arquivo existe
    if not os.path.exists(caminho_entrada):
        print(f"\n❌ ERRO: Arquivo não encontrado: {caminho_entrada}")
        return

    # Criar diretório de saída se não existir
    dir_saida = os.path.dirname(caminho_saida)
    if dir_saida and not os.path.exists(dir_saida):
        os.makedirs(dir_saida)

    # Deletar arquivo de saída se existir
    if os.path.exists(caminho_saida):
        print(f">>> Deletando arquivo existente: {caminho_saida}")
        os.remove(caminho_saida)

    # Processar
    try:
        resultados = processar_arquivo(
            caminho_entrada, caminho_saida, dados, todos_item
        )

        print("\n" + "=" * 60)
        print("RESUMO DO PROCESSAMENTO")
        print("=" * 60)
        print(
            f"Total fórmulas fator: {resultados['formulas_fator_comp'] + resultados['formulas_fator_aux']}"
        )
        print(f"Total hyperlinks: {resultados['hyperlinks_criados']}")

    except Exception as e:
        print(f"\n❌ ERRO durante o processamento: {str(e)}")
        import traceback

        traceback.print_exc()


if __name__ == "__main__":
    main()
