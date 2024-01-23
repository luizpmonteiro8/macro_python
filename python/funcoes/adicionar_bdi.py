from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.workbook import Workbook


def adicionar_bdi(workbook: Workbook, dados):
    # Obter valores
    planilha_fator = dados.get('planilhaFator', 'RESUMO')
    valor_bdi = dados.get('BDI', '28.82')
    valor_bdi_formatado = "{:.2%}".format(
        float(valor_bdi)/100).replace('.', ',')
    linha = dados.get('posicaoFator', {}).get('linha', 4) + 1
    colunaString = dados.get('posicaoFator', {}).get('coluna', "G")
    colunaNumber = column_index_from_string(colunaString)

    # Selecionar a planilha
    sheet_resumo = workbook[planilha_fator]

    # Selecionar a célula
    cell = sheet_resumo.cell(
        row=linha, column=colunaNumber)

    # Definir o valor da célula como "FATOR"
    cell.value = "BDI"

    # Selecionar a célula
    cell_value = sheet_resumo.cell(
        row=linha, column=colunaNumber+1)

    # Definir o valor da célula
    cell_value.value = valor_bdi_formatado

    # Adicionar a definição de nome "FATOR" referenciando a célula
    ref = f"{planilha_fator}!${get_column_letter(colunaNumber+1)}${linha}"
    definicao_nome = DefinedName(name="BDI", attr_text=ref)
    workbook.defined_names.add(definicao_nome)
