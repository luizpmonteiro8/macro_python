from openpyxl.utils import column_index_from_string

from funcoes.comum.buscar_palavras import (buscar_palavra_com_linha,
                                           buscar_palavra_com_linha_exato)
from funcoes.comum.copiar_coluna import copiar_coluna_com_numeros
from funcoes.comum.valor_bdi_final import valor_bdi_final


def copiar_colunas(sheet, dados):
    # Obter informações de coluna do JSON
    coluna_origem = dados.get(
        'colunaParaCopiarAux', {}).get('de', 'E')
    coluna_destino = dados.get(
        'colunaParaCopiarAux', {}).get('para', 'L')

    coluna_origem1 = dados.get(
        'colunaParaCopiarAux1', {}).get('de', 'F')
    coluna_destino1 = dados.get(
        'colunaParaCopiarAux1', {}).get('para', 'M')

    copiar_coluna_com_numeros(sheet, coluna_origem, coluna_destino)
    copiar_coluna_com_numeros(sheet, coluna_origem1, coluna_destino1)


def adicionar_formula_preco_unitario_menos_preco_antigo(sheet, dados):
    coluna_origem = dados.get(
        'colunaParaCopiarAux1', {}).get('para', 'M')
    coluna_preco_unitario = dados.get(
        'composicaoPrecoUnitario', 'F')
    linha_ini = 1
    final_linha = sheet.max_row + 1

    for x in range(linha_ini, final_linha):
        if sheet[f'{coluna_origem}{x}'].value is not None:
            coluna_destino = column_index_from_string(coluna_origem) + 1
            formula = f'=({coluna_origem}{x}-{coluna_preco_unitario}{x})'
            sheet.cell(row=x, column=coluna_destino).value = formula


def fator_nos_item_totais(sheet, dados,
                          linha_inicial_comp,
                          linha_final_comp,
                          nome,
                          totalNome,
                          coeficiente,
                          adicionar_fator,
                          ):
    coluna_descricao_composicao = dados.get(
        'colunaDescricaoComposicao', 'A'
    )
    coluna_totais_composicao = dados.get(
        'colunaTotaisComposicao', 'E'
    )
    coluna_totais_valor_composicao = dados.get(
        'colunaTotaisValorComposicao', 'G'
    )
    coluna_preco_unit = dados.get(
        "composicaoPrecoUnitario", "F"
    )
    coluna_coefieciente = dados.get(
        "composicaoCoeficiente", "E"
    )
    coluna_preco_unitario_antigo = dados.get(
        'colunaParaCopiarComposicao1', {}).get('para', 'M')
    coluna_coeficiente_antigo = dados.get(
        'colunaParaCopiarComposicao', {}).get('para', 'L')

    # verifica se tem material na composicao
    inicial = buscar_palavra_com_linha_exato(
        sheet, coluna_descricao_composicao,
        nome,
        linha_inicial_comp, linha_final_comp)
    final = buscar_palavra_com_linha_exato(
        sheet, coluna_totais_composicao,
        totalNome,
        linha_inicial_comp, linha_final_comp)

    if (inicial > -1 and final > -1
        and inicial < linha_final_comp
            and inicial > linha_inicial_comp):
        # total final
        soma_formula = (
            f'=SUM('
            f'{coluna_totais_valor_composicao}{inicial+1}:'
            f'{coluna_totais_valor_composicao}{final-1}'
            f')'
        )
        sheet[f'{coluna_totais_valor_composicao}{final}'].value = soma_formula
        for y in range(inicial+1, final):
            # total linha
            sheet[f'{coluna_totais_valor_composicao}{y}'].value = (
                f'=ROUND({coluna_coefieciente}{y}*{coluna_preco_unit}{y}, 2)'
            )
            if coeficiente and adicionar_fator:
                sheet[f'{coluna_coefieciente}{y}'].value = (
                    f'={coluna_coeficiente_antigo}{y}*FATOR'
                )
            else:
                if adicionar_fator:
                    sheet[f'{coluna_preco_unit}{y}'].value = (
                        f'=ROUND({coluna_preco_unitario_antigo}{y}*FATOR, 2)'
                    )

        return inicial, final


def buscar_auxiliar_no_aux(workbook, dados, linha, linha_total):
    # busca dentro de auxiliar os auxiliares
    sheet_name_aux = dados.get(
        'planilhaAuxiliar', 'COMPOSICOES AUXILIARES')
    sheet_planilha_aux = workbook[sheet_name_aux]

    coluna_item = dados.get(
        'colunaItemDescricaoComposicao', 'B')
    coluna_desc_aux = dados.get(
        'colunaDescricaoAuxiliar', 'A'
    )
    coluna_valor_com_dbi = dados.get(
        "colunaValorComBdi", "E",
    )
    coluna_valor = dados.get(
        "colunaValor", "G",
    )
    coluna_preco_aux = dados.get(
        'composicaoPrecoUnitario', 'F'
    )
    coluna_totais_composicao = dados.get(
        'colunaTotaisComposicao', 'E'
    )
    valor_string = dados.get(
        'valor', 'VALOR:'
    )
    ultima_linha = sheet_planilha_aux.max_row

    itens_array = []

    # Iterar sobre as chaves que começam com "item"
    for chave, valor in dados.items():
        if chave.startswith("item"):
            itens_array.append(valor)

    for x in range(linha, linha_total):
        cod = sheet_planilha_aux[f'{coluna_desc_aux}{x}'].value
        item = sheet_planilha_aux[f'{coluna_item}{x}'].value

        if item is not None:
            linha_inicial = buscar_palavra_com_linha(
                sheet_planilha_aux, coluna_desc_aux, cod + ' ' + item, 1,
                ultima_linha
            )

            linha_final = buscar_palavra_com_linha_exato(
                sheet_planilha_aux, coluna_valor_com_dbi,
                valor_string, linha_inicial, ultima_linha
            )

            # adicionando formula no preco unitario em auxiliar
            sheet_planilha_aux[f'{coluna_preco_aux}{x}'].value = (
                f'=\'{sheet_name_aux}\'!{coluna_valor}{linha_final}'
            )

            final_total_linha_array = []

            for item in itens_array:
                resultado_fator = fator_nos_item_totais(
                    sheet_planilha_aux, dados,
                    linha_inicial,
                    linha_final,
                    item['nome'], item['total'],
                    True if item['fatorCoeficiente'] == 'Sim' else False,
                    True if item['adicionarFator'] == 'Sim' else False
                )
                if resultado_fator is not None:
                    linha_desc, linha_total = resultado_fator
                if resultado_fator is not None and linha_total is not None:
                    final_total_linha_array.append(linha_total)

                if (item['buscarAuxiliar'] is not None
                        and item['buscarAuxiliar'] == 'Sim'
                        and resultado_fator is not None
                        and linha_desc > 0
                        and linha_total > 0):
                    buscar_auxiliar_no_aux(
                        workbook, dados, linha_desc, linha_total)

            # total no VALOR:
            if final_total_linha_array:
                linha_valor_sum = buscar_palavra_com_linha(
                    sheet_planilha_aux, coluna_totais_composicao, valor_string,
                    linha_inicial, linha_final+1)

                if linha_valor_sum > 0:
                    formula_soma = (
                        '=SUM(' +
                        ','.join([f'{coluna_valor}{linha}'
                                  for linha in final_total_linha_array]) +
                        ')'
                    )

                    # Atribui a fórmula à célula específica
                    sheet_planilha_aux[
                        f'{coluna_valor}{linha_valor_sum}'
                    ].value = formula_soma
                else:
                    print("A linha_valor_sum não é maior que zero.")


def adicionar_fator_totais_aux(workbook, dados, linhaIni, linhaFim):
    # chamado no adicionar_fator_comp
    sheet_name_aux = dados.get(
        'planilhaAuxiliar', 'COMPOSICOES AUXILIARES')
    sheet_planilha_aux = workbook[sheet_name_aux]

    coluna_totais_composicao = dados.get(
        'colunaTotaisComposicao', 'E'
    )

    valorString = dados.get(
        'valor', 'VALOR:')
    coluna_valor_string = dados.get(
        'colunaValor', 'G'
    )

    itens_array = []

    # Iterar sobre as chaves que começam com "item"
    for chave, valor in dados.items():
        if chave.startswith("item"):
            itens_array.append(valor)

    final_total_linha_array = []

    for item in itens_array:
        resultado_fator = fator_nos_item_totais(
            sheet_planilha_aux, dados,
            linhaIni,
            linhaFim,
            item['nome'], item['total'],
            True if item['fatorCoeficiente'] == 'Sim' else False,
            True if item['adicionarFator'] == 'Sim' else False
        )

        if resultado_fator is not None:
            linha_desc, linha_total = resultado_fator
        if (resultado_fator is not None and linha_total is not None):
            final_total_linha_array.append(linha_total)
        if (item['buscarAuxiliar'] is not None
                and item['buscarAuxiliar'] == 'Sim'
                and resultado_fator is not None
                and linha_desc > 0
                and linha_total > 0):
            buscar_auxiliar_no_aux(workbook, dados, linha_desc, linha_total)

        # total no VALOR:
        if final_total_linha_array:
            linha_valor_sum = buscar_palavra_com_linha(
                sheet_planilha_aux, coluna_totais_composicao, valorString,
                linhaIni, linhaFim+1)
            if linha_valor_sum > 0:
                formula_soma = (
                    '=SUM(' +
                    ','.join([f'{coluna_valor_string}{linha}'
                              for linha in final_total_linha_array]) +
                    ')')
                # Atribui a fórmula à célula específica
                sheet_planilha_aux[
                    f'{coluna_valor_string}{linha_valor_sum}'
                ].value = formula_soma
            else:
                print("A linha_valor_sum não é maior que zero.")


def adicionar_fator_aux(workbook, dados):
    sheet_name_aux = dados.get(
        'planilhaAuxiliar', 'COMPOSICOES AUXILIARES')
    sheet_planilha_aux = workbook[sheet_name_aux]

    copiar_colunas(sheet_planilha_aux, dados)
    adicionar_formula_preco_unitario_menos_preco_antigo(
        sheet_planilha_aux, dados
    )
    valor_bdi_final(sheet_planilha_aux, dados)
