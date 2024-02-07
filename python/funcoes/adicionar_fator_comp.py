import tkinter as tk

from openpyxl.utils import column_index_from_string

from funcoes.adicionar_fator_aux import adicionar_fator_totais_aux
from funcoes.comum.buscar_palavras import (buscar_palavra_com_linha,
                                           buscar_palavra_com_linha_exato)
from funcoes.comum.copiar_coluna import copiar_coluna_com_numeros
from funcoes.comum.valor_bdi_final import valor_bdi_final


def copiar_colunas(sheet, dados):
    # Obter informações de coluna do JSON
    coluna_origem = dados.get(
        'colunaParaCopiarComposicao', {}).get('de', 'E')
    coluna_destino = dados.get(
        'colunaParaCopiarComposicao', {}).get('para', 'L')

    coluna_origem1 = dados.get(
        'colunaParaCopiarComposicao1', {}).get('de', 'F')
    coluna_destino1 = dados.get(
        'colunaParaCopiarComposicao1', {}).get('para', 'M')

    copiar_coluna_com_numeros(sheet, coluna_origem, coluna_destino)
    copiar_coluna_com_numeros(sheet, coluna_origem1, coluna_destino1)


def adicionar_formula_preco_unitario_menos_preco_antigo(sheet, dados):
    coluna_origem = dados.get(
        'colunaParaCopiarComposicao1', {}).get('para', 'M')
    coluna_preco_unitario = dados.get(
        'composicaoPrecoUnitario', 'F')
    linha_ini = 1
    final_linha = sheet.max_row + 1

    for x in range(linha_ini, final_linha):
        if sheet[f'{coluna_origem}{x}'].value is not None:
            coluna_destino = column_index_from_string(coluna_origem) + 1
            formula = f'=({coluna_origem}{x}-{coluna_preco_unitario}{x})'
            sheet.cell(row=x, column=coluna_destino).value = formula


def fator_nos_item_totais(sheet, dados,
                          linha_inicial_comp,
                          linha_final_comp,
                          nome,
                          totalNome,
                          coeficiente,
                          adicionar_fator):
    coluna_descricao_composicao = dados.get(
        'colunaDescricaoComposicao', 'A'
    )
    coluna_totais_composicao = dados.get(
        'colunaTotaisComposicao', 'E'
    )
    coluna_totais_valor_composicao = dados.get(
        'colunaTotaisValorComposicao', 'G'
    )
    coluna_preco_unit = dados.get(
        "composicaoPrecoUnitario", "F"
    )
    coluna_coefieciente = dados.get(
        "composicaoCoeficiente", "E"
    )
    coluna_preco_unitario_antigo = dados.get(
        'colunaParaCopiarComposicao1', {}).get('para', 'M')
    coluna_coeficiente_antigo = dados.get(
        'colunaParaCopiarComposicao', {}).get('para', 'L')

    # verifica se tem material na composicao
    inicial = buscar_palavra_com_linha_exato(
        sheet, coluna_descricao_composicao,
        nome,
        linha_inicial_comp, linha_final_comp)
    final = buscar_palavra_com_linha_exato(
        sheet, coluna_totais_composicao,
        totalNome,
        linha_inicial_comp, linha_final_comp)

    if (inicial > -1 and final > -1
        and inicial < linha_final_comp
            and inicial > linha_inicial_comp):
        # total final
        soma_formula = (
            f'=SUM('
            f'{coluna_totais_valor_composicao}{inicial+1}:'
            f'{coluna_totais_valor_composicao}{final-1}'
            f')'
        )
        sheet[f'{coluna_totais_valor_composicao}{final}'].value = soma_formula
        for y in range(inicial+1, final):
            # total linha
            sheet[f'{coluna_totais_valor_composicao}{y}'].value = (
                f'=ROUND({coluna_coefieciente}{y}*{coluna_preco_unit}{y}, 2)'
            )
            if coeficiente and adicionar_fator:
                sheet[f'{coluna_coefieciente}{y}'].value = (
                    f'={coluna_coeficiente_antigo}{y}*FATOR'
                )
            else:
                if adicionar_fator:
                    sheet[f'{coluna_preco_unit}{y}'].value = (
                        f'=ROUND({coluna_preco_unitario_antigo}{y}*FATOR, 2)'
                    )

        return inicial, final


def buscar_comp_auxiliar(workbook, dados, linha, linha_total):
    sheet_name_comp = dados.get(
        'planilhaComposicao', 'COMPOSICAO')
    sheet_planilha_comp = workbook[sheet_name_comp]
    sheet_name_aux = dados.get(
        'planilhaAuxiliar', 'COMPOSICOES AUXILIARES')
    sheet_planilha_aux = workbook[sheet_name_aux]

    coluna_item = dados.get(
        'colunaItemDescricaoComposicao', 'B')
    coluna_desc_aux = dados.get(
        'colunaDescricaoAuxiliar', 'A'
    )
    coluna_valor_com_dbi = dados.get(
        "colunaValorComBdi", "E",
    )
    coluna_valor = dados.get(
        "colunaValor", "G",
    )
    coluna_preco_unit = dados.get(
        "composicaoPrecoUnitario", "F"
    )
    valor_string = dados.get(
        'valor', 'VALOR:'
    )

    ultima_linha_busca = 1
    ultima_linha_aux = sheet_planilha_aux.max_row

    # noma para busca na auxiliar
    for x in range(linha, linha_total):
        cod = sheet_planilha_comp[f'{coluna_desc_aux}{x}'].value
        nome = sheet_planilha_comp[f'{coluna_item}{x}'].value

        if nome is not None:
            linha_inicial = buscar_palavra_com_linha(
                sheet_planilha_aux, coluna_desc_aux, cod + ' ' + nome,
                ultima_linha_busca, ultima_linha_aux)
            if linha_inicial > -1:
                linha_final = buscar_palavra_com_linha(
                    sheet_planilha_aux, coluna_valor_com_dbi, valor_string,
                    linha_inicial, ultima_linha_aux
                )
                # adiciona fator e totais no auxiliar
                adicionar_fator_totais_aux(
                    workbook, dados, linha_inicial, linha_final
                )
                sheet_planilha_comp[f'{coluna_preco_unit}{x}'].value = (
                    f'=\'{sheet_name_aux}\'!{coluna_valor}{linha_final}')


def adicionar_fator_totais(workbook, dados, linhaIni, linhaFim):
    sheet_name = dados.get(
        'planilha', 'PLANILHA ORCAMENTARIA')
    sheet_planilha = workbook[sheet_name]
    sheet_name_comp = dados.get(
        'planilhaComposicao', 'COMPOSICAO')
    sheet_planilha_comp = workbook[sheet_name_comp]
    sheet_comp_linha_fim = sheet_planilha_comp.max_row + 1

    coluna_preco_planilha = dados.get(
        'planilhaPrecoUnitario', 'G'
    )
    coluna_descricao_composicao = dados.get(
        'colunaDescricaoComposicao', 'A'
    )
    coluna_totais_composicao = dados.get(
        'colunaTotaisComposicao', 'E'
    )

    #
    valor_com_bdi = dados.get(
        'valorComBdi', 'VALOR COM BDI:')
    valorString = dados.get(
        'valor', 'VALOR:')
    coluna_cod = dados.get(
        'colunaCodEmPlanilhaParaBuscaEmComposicao', 'B'
    )
    coluna_descricao = dados.get(
        'colunaDescricaoEmPlanilhaParaBuscaEmComposicao', 'C'
    )
    coluna_valor_bdi = dados.get(
        'colunaValorComBdi', 'E'
    )
    coluna_valor_string = dados.get(
        'colunaValor', 'G'
    )

    itens_array = []

    # Iterar sobre as chaves que começam com "item"
    for chave, valor in dados.items():
        if chave.startswith("item"):
            itens_array.append(valor)

    # evitar usar valor errado iniciando no ultimo que foi buscado
    linha_final_iniciar_busca = 1

    for x in range(linhaIni, linhaFim):
        # busca nome da descricao na planilha
        cod = sheet_planilha[f'{coluna_cod}{x}'].value
        coluna_busca_value = sheet_planilha[f'{coluna_descricao}{x}'].value

        if coluna_busca_value is not None:
            # busca nome da descricao na composicao
            linha_inicial_comp = -1
            linha_inicial_comp = buscar_palavra_com_linha(
                sheet_planilha_comp, coluna_descricao_composicao,
                cod + ' ' + coluna_busca_value, linha_final_iniciar_busca,
                sheet_comp_linha_fim)
            if (linha_inicial_comp == -1):
                linha_inicial_comp = buscar_palavra_com_linha(
                    sheet_planilha_comp, coluna_descricao_composicao,
                    cod, linha_final_iniciar_busca,
                    sheet_comp_linha_fim)

            if (linha_inicial_comp == -1):
                tk.messagebox.showwarning(
                    "Aviso",
                    "Não foi encontrado o item na composição. " +
                    cod + " " + coluna_busca_value +
                    "Verifique se o item existe em composição. ")

            # busca linha final pelo valor bdi
            linha_final_comp = buscar_palavra_com_linha(
                sheet_planilha_comp, coluna_valor_bdi, valor_com_bdi,
                linha_inicial_comp, sheet_comp_linha_fim)
            # linha por onde ele vai buscar inicialmente na proxima rodada
            linha_final_iniciar_busca = linha_final_comp

            # adicionando formula no preco unitario em planilha
            sheet_planilha[f'{coluna_preco_planilha}{x}'].value = (
                f'={sheet_name_comp}!{coluna_valor_string}{linha_final_comp}')

            final_total_linha_array = []

            for item in itens_array:
                resultado_fator = fator_nos_item_totais(
                    sheet_planilha_comp, dados,
                    linha_inicial_comp,
                    linha_final_comp,
                    item['nome'],
                    item['total'],
                    True if item['fatorCoeficiente'] == 'Sim' else False,
                    True if item['adicionarFator'] == 'Sim' else False,
                )
                if resultado_fator is not None:
                    linha_desc, linha_total = resultado_fator
                if resultado_fator is not None and linha_total is not None:
                    final_total_linha_array.append(linha_total)
                # busca auxiliar
                if (item['buscarAuxiliar'] is not None
                        and item['buscarAuxiliar'] == 'Sim'
                        and resultado_fator is not None
                        and linha_desc > 0
                        and linha_total > 0):
                    buscar_comp_auxiliar(
                        workbook, dados, linha_desc, linha_total)

            # total no VALOR:
            if final_total_linha_array:
                linha_valor_sum = buscar_palavra_com_linha(
                    sheet_planilha_comp, coluna_totais_composicao, valorString,
                    linha_inicial_comp, linha_final_comp)

                if linha_valor_sum > 0:
                    formula_soma = (
                        '=SUM(' +
                        ','.join([f'{coluna_valor_string}{linha}'
                                  for linha in final_total_linha_array]) +
                        ')'
                    )

                    # Atribui a fórmula à célula específica
                    sheet_planilha_comp[
                        f'{coluna_valor_string}{linha_valor_sum}'
                    ].value = formula_soma
                else:
                    print("A linha_valor_sum não é maior que zero.")


def adicionar_fator_comp(workbook, dados, linhaIniPlan, linhaFimPlan):
    sheet_name_comp = dados.get(
        'planilhaComposicao', 'COMPOSICOES')
    sheet_planilha_comp = workbook[sheet_name_comp]

    copiar_colunas(sheet_planilha_comp, dados)
    adicionar_formula_preco_unitario_menos_preco_antigo(
        sheet_planilha_comp, dados)
    adicionar_fator_totais(workbook, dados, linhaIniPlan, linhaFimPlan)
    valor_bdi_final(sheet_planilha_comp, dados)
