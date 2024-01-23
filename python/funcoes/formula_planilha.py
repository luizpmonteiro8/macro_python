from copy import copy

from openpyxl import Workbook
from openpyxl.utils import column_index_from_string

from funcoes.comum.copiar_coluna import copiar_coluna


def copiar_coluna_planilha(sheet, dados):
    # Obter informações de coluna do JSON
    coluna_origem = dados.get(
        'colunaParaCopiar', {}).get('de', 'G')
    coluna_destino = dados.get(
        'colunaParaCopiar', {}).get('para', 'K')

    copiar_coluna(sheet, coluna_origem, coluna_destino)


def copiar_estilo(origem, destino):
    destino.font = copy(origem.font)
    destino.border = copy(origem.border)
    destino.fill = copy(origem.fill)
    destino.number_format = copy(origem.number_format)
    destino.protection = copy(origem.protection)
    destino.alignment = copy(origem.alignment)


def formula_planilha(woorBook: Workbook, linhaIni, linhaFin, dados):
    # Obter informações de coluna do JSON
    planilha = dados.get('planilha', 'PLANILHA ORCAMENTARIA')

    quantidade_string = dados.get('planilhaQuantidade', 'F')
    preco_unitario_string = dados.get('planilhaPrecoUnitario', 'G')

    coluna_preco_total = dados.get('planilhaPrecoTotal', 'H')
    coluna_preco_total_number = column_index_from_string(coluna_preco_total)

    coluna_value_string = dados.get('colunaParaCopiar', {}).get('para', 'K')
    coluna_value_number = column_index_from_string(coluna_value_string)

    ws = woorBook[planilha]

    # Iterar sobre as linhas
    for x in range(linhaIni, linhaFin):
        # Verificar se o valor na coluna 11 não está vazio ou seja não é total
        if ws.cell(row=x, column=coluna_value_number).value is not None:
            # Calcular e atribuir o valor na coluna
            ws.cell(row=x, column=coluna_value_number+1).value = (
                f"={preco_unitario_string}{x}-"
                f"{coluna_value_string}{x}"
            )

            # Copiar o estilo da célula de quantidade para preço total
            copiar_estilo(ws.cell(row=x, column=column_index_from_string(
                quantidade_string)), ws.cell(row=x,
                                             column=coluna_preco_total_number))

            # Calcular e atribuir o valor na coluna de preço total
            ws.cell(row=x, column=coluna_preco_total_number).value = (
                f"=ROUND({quantidade_string}{x}*{preco_unitario_string}{x}, 2)"
            )
