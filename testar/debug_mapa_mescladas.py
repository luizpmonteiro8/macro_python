"""Script para testar a função _construir_mapa_mescladas com mais detalhes"""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from openpyxl.cell.cell import MergedCell

# Importar a função do módulo verificar_auxiliar_fator
from funcoes.planilha.funcoes.verificar_auxiliar_fator import _construir_mapa_mescladas

# Abrir o arquivo Excel
caminho_arquivo = "testar/excel-final/orcamento-completo22-04-2026_08-04-50.xlsx"
workbook = openpyxl.load_workbook(caminho_arquivo, data_only=False)

# Listar todas as planilhas
print("=== PLANILHAS NO ARQUIVO ===")
for nome in workbook.sheetnames:
    print(f"  - {nome}")

# Testar na planilha COMPOSICOES AUXILIARES
planilha_aux = "COMPOSICOES AUXILIARES"
col_desc = 1  # Coluna A

if planilha_aux in workbook.sheetnames:
    sheet = workbook[planilha_aux]

    print(f"\n=== CHAMANDO _construir_mapa_mescladas ===")
    mapa_mescladas = _construir_mapa_mescladas(sheet, col_desc)

    print(f"\nTotal de títulos encontrados no mapa: {len(mapa_mescladas)}")

    # Mostrar todos os códigos encontrados
    print("\n=== CÓDIGOS ENCONTRADOS NO MAPA ===")
    for codigo, linha in sorted(mapa_mescladas.items(), key=lambda x: x[1]):
        print(f"  Linha {linha}: {codigo}")

    # ==========================================
    # BUSCAR PELO CÓDIGO ESPECÍFICO
    # 88248 AUXILIAR DE ENCANADOR OU BOMBEIRO HIDRÁULICO COM ENCARGOS COMPLEMENTARES
    # ==========================================
    print("\n" + "=" * 80)
    print("=== BUSCANDO PELO CÓDIGO: 88248 AUXILIAR DE ENCANADOR... ===")
    print("=" * 80)

    # Código procurado
    codigo_procurado = "88248"

    # Verificar se existe no mapa
    encontrado_no_mapa = False
    for codigo, linha in mapa_mescladas.items():
        if codigo_procurado in codigo:
            encontrado_no_mapa = True
            print(f"\n[OK] ENCONTRADO NO MAPA!")
            print(f"  Código: {codigo}")
            print(f"  Linha: {linha}")

            # Mostrar a célula na planilha (limpando caracteres especiais)
            cell = sheet.cell(row=linha, column=col_desc)
            val_limpo = (
                str(cell.value).replace("\u200b", "").replace("\ufeff", "").strip()
                if cell.value
                else ""
            )
            print(f"  Valor completo: {val_limpo}")

    if not encontrado_no_mapa:
        print(
            f"\n[X] CODIGO '{codigo_procurado}' NAO ENCONTRADO NO MAPA DE CELULAS MESCLADAS"
        )

        # Vamos procurar diretamente na planilha
        print("\n=== PROCURANDO DIRETAMENTE NA PLANILHA ===")
        max_row = min(sheet.max_row, 20000)

        for linha in range(1, max_row + 1):
            cell = sheet.cell(row=linha, column=col_desc)

            # Pular MergedCells que não são master
            if isinstance(cell, MergedCell):
                continue

            valor = cell.value
            if valor and codigo_procurado in str(valor):
                print(f"\n[OK] ENCONTRADO NA LINHA {linha}!")
                print(f"  Valor: {valor}")

                # Verificar se está em uma célula mesclada
                for mr in sheet.merged_cells.ranges:
                    if (
                        mr.min_row <= linha <= mr.max_row
                        and mr.min_col <= col_desc <= mr.max_col
                    ):
                        print(f"  Faz parte de range mesclado: {mr}")
                        print(
                            f"  Tamanho: {(mr.max_row - mr.min_row + 1) * (mr.max_col - mr.min_col + 1)} células"
                        )
                        break

                # Mostrar contexto (próximas linhas)
                print("\n  Contexto (linhas seguintes):")
                for i in range(linha + 1, min(linha + 10, max_row + 1)):
                    ctx = sheet.cell(row=i, column=col_desc).value
                    if ctx:
                        print(f"    Linha {i}: {str(ctx)[:100]}")

                # Mostrar contexto (linhas anteriores)
                print("\n  Contexto (linhas anteriores):")
                for i in range(max(1, linha - 5), linha):
                    ctx = sheet.cell(row=i, column=col_desc).value
                    if ctx:
                        print(f"    Linha {i}: {str(ctx)[:100]}")

                break

# Verificar se a célula mesclada ao redor do código tem tamanho adequado
print("\n=== VERIFICANDO CÉLULAS MESCLADAS PRÓXIMAS AO CÓDIGO 88248 ===")
if planilha_aux in workbook.sheetnames:
    sheet = workbook[planilha_aux]
    col_desc = 1

    # Procurar todas as células mescladas e ver quais contêm ou estão próximas ao código
    for mr in sheet.merged_cells.ranges:
        if mr.min_col <= col_desc <= mr.max_col:
            total_celulas = (mr.max_row - mr.min_row + 1) * (
                mr.max_col - mr.min_col + 1
            )

            # Verificar se o tamanho é <= 3 (caso não entre no mapa)
            if total_celulas <= 3:
                val = sheet.cell(row=mr.min_row, column=col_desc).value
                if val and "88248" in str(val):
                    print(f"\n  Range: {mr}")
                    print(
                        f"  Tamanho: {total_celulas} células (<=3, por isso NÃO entra no mapa)"
                    )
                    print(f"  Valor: {val}")

workbook.close()
print("\n=== FIM DO TESTE ===")
