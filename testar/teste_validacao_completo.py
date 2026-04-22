"""
Teste de validação completo para verificar_auxiliar_fator.py
Simula caso real e valida cada seção do Excel
"""

import sys
import os

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)) + "/..")

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
import json
import copy

# Importar a função original
import importlib.util

spec = importlib.util.spec_from_file_location(
    "verificar_auxiliar_fator",
    os.path.dirname(os.path.abspath(__file__))
    + "/../funcoes/planilha/funcoes/verificar_auxiliar_fator.py",
)
mod = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mod)

verificar_auxiliar_fator = mod.verificar_auxiliar_fator


def limpar_celula(val):
    """Remove caracteres especiais"""
    if val is None:
        return ""
    return str(val).replace("\u200b", "").replace("\ufeff", "").strip()


def criar_workbook_teste(caminho_excel, caminho_json):
    """Cria uma cópia do workbook para teste"""
    # Carregar config
    with open(caminho_json, "r", encoding="utf-8") as f:
        dados = json.load(f)

    # Carregar Excel
    wb = load_workbook(caminho_excel)

    return wb, dados


def validar_resultado(wb, sheet_name, secoes_config, descricao_teste):
    """Valida os resultados de um teste"""
    sheet = wb[sheet_name]
    col_a = column_index_from_string("A")
    col_e = column_index_from_string("E")
    col_f = column_index_from_string("F")
    col_m = column_index_from_string("M")

    resultados = {
        "descricao": descricao_teste,
        "sucesso": True,
        "linhas_processadas": [],
        "erros": [],
    }

    for secao_info in secoes_config:
        nome_secao = secao_info["nome"]
        linhas = secao_info["linhas"]
        tipo = secao_info["tipo"]  # "preco" ou "coef"
        coluna_alvo = col_f if tipo == "preco" else col_e
        nome_coluna = "F (Preço Unitário)" if tipo == "preco" else "E (Coeficiente)"

        print(f"\n  → Validando {nome_secao} ({nome_coluna}):")

        for linha in linhas:
            desc = limpar_celula(sheet.cell(row=linha, column=col_a).value)
            valor_col = sheet.cell(row=linha, column=coluna_alvo).value
            valor_m = sheet.cell(row=linha, column=col_m).value

            # Verificar se tem fórmula de fator
            if (
                valor_col
                and isinstance(valor_col, str)
                and "*FATOR" in valor_col.upper()
            ):
                print(f"    ✓ Linha {linha}: '{desc[:30]}...' → {valor_col}")
                resultados["linhas_processadas"].append(linha)
            elif valor_col and isinstance(valor_col, str) and valor_col.startswith("="):
                print(f"    ✓ Linha {linha}: '{desc[:30]}...' → {valor_col}")
                resultados["linhas_processadas"].append(linha)
            elif not desc or desc.startswith("1."):
                # Descrição de composição, não é item
                pass
            else:
                print(
                    f"    ✗ Linha {linha}: '{desc[:30]}...' → Sem fator (valor: {valor_col})"
                )
                if secao_info.get("esperado", True):
                    resultados["erros"].append(f"{nome_secao} linha {linha} sem fator")
                    resultados["sucesso"] = False

    return resultados


def main():
    print("=" * 70)
    print("TESTE DE VALIDAÇÃO COMPLETO - verificar_auxiliar_fator.py")
    print("=" * 70)

    # Caminhos dos arquivos
    caminho_excel = r"testar\orcamento-completo.xlsx"
    caminho_json = r"config\valores_item.json"

    # Verificar se arquivos existem
    if not os.path.exists(caminho_excel):
        print(f"ERRO: Arquivo Excel não encontrado: {caminho_excel}")
        return

    if not os.path.exists(caminho_json):
        print(f"ERRO: Arquivo JSON não encontrado: {caminho_json}")
        return

    # Carregar arquivos
    print(f"\nCarregando Excel: {caminho_excel}")
    print(f"Carregando JSON: {caminho_json}")

    wb = load_workbook(caminho_excel)

    with open(caminho_json, "r", encoding="utf-8") as f:
        dados = json.load(f)

    # Mostrar config
    print("\n" + "=" * 70)
    print("CONFIGURAÇÃO CARREGADA (valores_item.json)")
    print("=" * 70)

    dados_itens = dados[0]

    print("\n--- ITENS COM FATOR ---")
    itens_fator = []
    for key, item in dados_itens.items():
        if not key.startswith("item") or not isinstance(item, dict):
            continue
        if item.get("adicionarFator") == "Sim":
            itens_fator.append(item)
            print(
                f"  • {item.get('nome')}: fatorCoef={item.get('fatorCoeficiente')}, "
                f"iniciaPor='{item.get('iniciaPor')}', naoIniciaPor='{item.get('naoIniciaPor')}'"
            )

    print("\n--- ITENS AUXILIARES ---")
    for key, item in dados_itens.items():
        if not key.startswith("item") or not isinstance(item, dict):
            continue
        if item.get("buscarAuxiliar") == "Sim" and item.get("adicionarFator") != "Sim":
            print(
                f"  • {item.get('nome')}: iniciaPor='{item.get('iniciaPor')}', "
                f"naoIniciaPor='{item.get('naoIniciaPor')}'"
            )

    # Criar cópia do workbook para teste
    print("\n" + "=" * 70)
    print("EXECUTANDO verificar_auxiliar_fator()")
    print("=" * 70)

    wb_teste = load_workbook(caminho_excel)

    # Executar a função
    resultado = verificar_auxiliar_fator(wb_teste, dados)

    print("\n--- RESULTADO DA EXECUÇÃO ---")
    print(f"  Fórmulas fator (COMPOSICOES): {resultado['formulas_fator_comp']}")
    print(f"  Fórmulas fator (AUXILIARES): {resultado['formulas_fator_aux']}")
    print(
        f"  Fórmulas auxiliares (COMPOSICOES): {resultado['formulas_auxiliares_comp']}"
    )
    print(f"  Fórmulas auxiliares (AUXILIARES): {resultado['formulas_auxiliares_aux']}")
    print(f"  Hyperlinks criados: {resultado['hyperlinks_criados']}")

    # Validar seções específicas
    print("\n" + "=" * 70)
    print("VALIDAÇÃO DE SEÇÕES ESPECÍFICAS")
    print("=" * 70)

    # COMPOSICOES - Material (linhas 24-33)
    print("\n### COMPOSICOES - Material (linhas 24-33) ###")
    sheet_comp = wb_teste["COMPOSICOES"]
    col_a = column_index_from_string("A")
    col_e = column_index_from_string("E")
    col_f = column_index_from_string("F")
    col_m = column_index_from_string("M")

    erros_material = []
    linhas_com_fator_material = []

    for row in range(24, 34):
        desc = limpar_celula(sheet_comp.cell(row=row, column=col_a).value)
        valor_f = sheet_comp.cell(row=row, column=col_f).value
        valor_m = sheet_comp.cell(row=row, column=col_m).value

        if desc and not desc.startswith("1.") and not desc.startswith("TOTAL"):
            if (
                valor_f
                and isinstance(valor_f, str)
                and ("*FATOR" in valor_f.upper() or valor_f.startswith("="))
            ):
                linhas_com_fator_material.append(row)
                print(f"  ✓ Linha {row}: '{desc}' → {valor_f}")
            else:
                if valor_m and isinstance(valor_m, (int, float)) and valor_m > 0:
                    erros_material.append(
                        f"Linha {row}: desc='{desc}', preco_coluna_F='{valor_f}', preco_coluna_M={valor_m}"
                    )
                    print(f"  ✗ Linha {row}: '{desc}' → F={valor_f}, M={valor_m}")

    # COMPOSICOES - Encargos Complementares
    print("\n### COMPOSICOES - Encargos Complementares ###")
    secoes_encontradas = []

    for row in range(1, 1500):
        val = limpar_celula(sheet_comp.cell(row=row, column=col_a).value)
        if val and val.upper() == "ENCARGOS COMPLEMENTARES":
            secoes_encontradas.append(("COMPOSICOES", row))

    erros_encargos = []
    linhas_com_fator_encargos = []

    for nome_sheet, linha_inicio in secoes_encontradas[:2]:
        print(f"\n  Seção encontrada na linha {linha_inicio}:")
        sheet = wb_teste[nome_sheet]

        # Encontrar próxima seção
        proxima_secao = None
        for row in range(linha_inicio + 1, min(linha_inicio + 100, sheet.max_row + 1)):
            val = limpar_celula(sheet.cell(row=row, column=col_a).value)
            if val and val.upper() in ["MATERIAL", "MÃO DE OBRA", "EQUIPAMENTO"]:
                proxima_secao = row
                break

        fim = (
            proxima_secao - 1
            if proxima_secao
            else min(linha_inicio + 50, sheet.max_row)
        )

        for row in range(linha_inicio + 1, fim + 1):
            desc = limpar_celula(sheet.cell(row=row, column=col_a).value)
            valor_f = sheet.cell(row=row, column=col_f).value
            valor_m = sheet.cell(row=row, column=col_m).value

            if desc and len(desc) > 3 and not desc.startswith("1."):
                if (
                    valor_f
                    and isinstance(valor_f, str)
                    and ("*FATOR" in valor_f.upper() or valor_f.startswith("="))
                ):
                    linhas_com_fator_encargos.append(row)
                    print(f"    ✓ Linha {row}: '{desc[:25]}' → {valor_f}")
                else:
                    if valor_m and isinstance(valor_m, (int, float)) and valor_m > 0:
                        erros_encargos.append(
                            f"Linha {row}: desc='{desc}', F='{valor_f}', M={valor_m}"
                        )
                        print(
                            f"    ✗ Linha {row}: '{desc[:25]}' → F={valor_f}, M={valor_m}"
                        )

    # COMPOSICOES - Mão de Obra
    print("\n### COMPOSICOES - Mão de Obra ###")
    secoes_mao = []

    for row in range(1, 1500):
        val = limpar_celula(sheet_comp.cell(row=row, column=col_a).value)
        if val and val.upper() == "MÃO DE OBRA":
            secoes_mao.append(("COMPOSICOES", row))

    erros_mao = []
    linhas_com_fator_mao = []

    for nome_sheet, linha_inicio in secoes_mao[:2]:
        print(f"\n  Seção encontrada na linha {linha_inicio}:")
        sheet = wb_teste[nome_sheet]

        # Encontrar próxima seção
        proxima_secao = None
        for row in range(linha_inicio + 1, min(linha_inicio + 100, sheet.max_row + 1)):
            val = limpar_celula(sheet.cell(row=row, column=col_a).value)
            if val and val.upper() in [
                "MATERIAL",
                "ENCARGOS COMPLEMENTARES",
                "EQUIPAMENTO",
            ]:
                proxima_secao = row
                break

        fim = (
            proxima_secao - 1
            if proxima_secao
            else min(linha_inicio + 50, sheet.max_row)
        )

        for row in range(linha_inicio + 1, fim + 1):
            desc = limpar_celula(sheet.cell(row=row, column=col_a).value)
            valor_e = sheet.cell(row=row, column=col_e).value
            valor_m = sheet.cell(row=row, column=col_m).value

            if desc and len(desc) > 3 and not desc.startswith("1."):
                if valor_e and isinstance(valor_e, str) and "*FATOR" in valor_e.upper():
                    linhas_com_fator_mao.append(row)
                    print(f"    ✓ Linha {row}: '{desc[:25]}' → {valor_e}")
                else:
                    if valor_m and isinstance(valor_m, (int, float)) and valor_m > 0:
                        erros_mao.append(
                            f"Linha {row}: desc='{desc}', E='{valor_e}', M={valor_m}"
                        )
                        print(
                            f"    ✗ Linha {row}: '{desc[:25]}' → E={valor_e}, M={valor_m}"
                        )

    # COMPOSICOES AUXILIARES
    print("\n### COMPOSICOES AUXILIARES - Encargos/Mão de Obra ###")
    sheet_aux = wb_teste["COMPOSICOES AUXILIARES"]

    erros_aux = []
    linhas_com_fator_aux = []

    # Verificar linhas 639-656
    for row in range(639, 657):
        desc = limpar_celula(sheet_aux.cell(row=row, column=col_a).value)
        valor_f = sheet_aux.cell(row=row, column=col_f).value
        valor_m = sheet_aux.cell(row=row, column=col_m).value

        if desc and len(desc) > 3:
            if (
                valor_f
                and isinstance(valor_f, str)
                and ("*FATOR" in valor_f.upper() or valor_f.startswith("="))
            ):
                linhas_com_fator_aux.append(row)
                print(f"  ✓ Linha {row}: '{desc[:25]}' → {valor_f}")
            else:
                if valor_m and isinstance(valor_m, (int, float)) and valor_m > 0:
                    erros_aux.append(
                        f"Linha {row}: desc='{desc}', F='{valor_f}', M={valor_m}"
                    )
                    print(f"  ✗ Linha {row}: '{desc[:25]}' → F={valor_f}, M={valor_m}")

    # Resumo final
    print("\n" + "=" * 70)
    print("RESUMO FINAL")
    print("=" * 70)

    print(f"\n  MATERIAL (linhas 24-33):")
    print(f"    Linhas com fator: {len(linhas_com_fator_material)}")
    print(f"    Erros (sem fator): {len(erros_material)}")

    print(f"\n  ENCARGOS COMPLEMENTARES:")
    print(f"    Linhas com fator: {len(linhas_com_fator_encargos)}")
    print(f"    Erros (sem fator): {len(erros_encargos)}")

    print(f"\n  MÃO DE OBRA (coeficiente):")
    print(f"    Linhas com fator: {len(linhas_com_fator_mao)}")
    print(f"    Erros (sem fator): {len(erros_mao)}")

    print(f"\n  COMPOSICOES AUXILIARES (linhas 639-656):")
    print(f"    Linhas com fator: {len(linhas_com_fator_aux)}")
    print(f"    Erros (sem fator): {len(erros_aux)}")

    total_erros = (
        len(erros_material) + len(erros_encargos) + len(erros_mao) + len(erros_aux)
    )
    total_fator = (
        len(linhas_com_fator_material)
        + len(linhas_com_fator_encargos)
        + len(linhas_com_fator_mao)
        + len(linhas_com_fator_aux)
    )

    print(f"\n  TOTAL DE LINHAS COM FATOR: {total_fator}")
    print(f"  TOTAL DE ERROS: {total_erros}")

    if total_erros == 0:
        print("\n  ✓✓✓ TESTE PASSOU - TODAS AS SEÇÕES PROCESSADAS CORRETAMENTE ✓✓✓")
    else:
        print(f"\n  ✗✗✗ TESTE FALHOU - {total_erros} ERROS ENCONTRADOS ✗✗✗")
        print("\n  Linhas com problema:")
        for erro in erros_material[:5]:
            print(f"    Material: {erro}")
        for erro in erros_encargos[:5]:
            print(f"    Encargos: {erro}")
        for erro in erros_mao[:5]:
            print(f"    Mão de Obra: {erro}")
        for erro in erros_aux[:5]:
            print(f"    Auxiliares: {erro}")

    # Salvar resultado
    resultado_teste = {
        "sucesso": total_erros == 0,
        "totais": {"linhas_com_fator": total_fator, "erros": total_erros},
        "material": {
            "com_fator": len(linhas_com_fator_material),
            "erros": erros_material,
        },
        "encargos": {
            "com_fator": len(linhas_com_fator_encargos),
            "erros": erros_encargos,
        },
        "mao_de_obra": {"com_fator": len(linhas_com_fator_mao), "erros": erros_mao},
        "auxiliares": {"com_fator": len(linhas_com_fator_aux), "erros": erros_aux},
    }

    # Salvar JSON com resultado
    with open(r"testar\resultado_teste.json", "w", encoding="utf-8") as f:
        json.dump(resultado_teste, f, indent=2, ensure_ascii=False)

    print(f"\n  Resultado salvo em: testar/resultado_teste.json")


if __name__ == "__main__":
    main()
