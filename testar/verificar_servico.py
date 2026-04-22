"""Verificar se a seção Servico NAO tem fator aplicado."""

from openpyxl import load_workbook

wb = load_workbook(r"testar\excel-final\orcamento-processado-final.xlsx")
sheet = wb["COMPOSICOES"]

print("=== Verificando linhas 68-85 (area de Mao de Obra e Servico) ===")
for row in range(68, 86):
    col_a = sheet.cell(row=row, column=1).value
    col_e = sheet.cell(row=row, column=5).value
    col_f = sheet.cell(row=row, column=6).value

    tem_fator_e = "*FATOR" in str(col_e).upper() if col_e else False
    tem_fator_f = "*FATOR" in str(col_f).upper() if col_f else False

    status_e = "OK" if not tem_fator_e else "TEM FATOR!"
    status_f = "OK" if not tem_fator_f else "TEM FATOR!"

    if col_a and ("MAO" in str(col_a).upper() or "SERVICO" in str(col_a).upper()):
        print(f"\n>>> LINHA {row}: {col_a}")
        print(f"    Coef (E): {col_e} - {status_e}")
        print(f"    Preco (F): {col_f} - {status_f}")
    elif col_a:
        print(f"\nLinha {row}: {col_a}")
        print(f"    Coef (E): {col_e} - {status_e}")
        print(f"    Preco (F): {col_f} - {status_f}")
